from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from harness.config import HarnessConfig
from harness.verification import validate_workflow
from projects.operaRezervasyon.workflow import (
    OperaRezervasyonFromExcelWorkflow,
    login_reached_authenticated_state,
    missing_credential_names,
    read_reservation_records,
    validate_reservation_fields,
)


def test_opera_rezervasyon_yaml_is_valid_and_uses_input_references():
    workflow_path = Path("projects/operaRezervasyon/workflows/main.yaml")
    workflow = yaml.safe_load(workflow_path.read_text(encoding="utf-8"))

    assert validate_workflow(workflow) == []
    assert "credentials" not in workflow
    assert workflow["steps"][0]["success_check"][0]["value"] == "${inputs.input_excel}"
    assert workflow["steps"][1]["action"]["path"] == "${inputs.input_excel}"
    assert workflow["steps"][1]["success_check"][0]["value"] == "${inputs.sheet}"
    payload = json.dumps(workflow)
    assert "KRONOVITAAI" not in payload
    assert "Nt~" not in payload


def test_opera_rezervasyon_config_defaults_to_dry_run_validation():
    config = HarnessConfig.from_yaml("projects/operaRezervasyon/config.yaml")

    assert config.variables["opera_live_preflight"] is False
    assert config.variables["opera_allow_login"] is False


def test_login_preflight_does_not_pass_when_login_form_remains():
    assert not login_reached_authenticated_state(
        current_url="https://example.test/ITOAS/operacloud/login",
        password_visible=True,
        body_text="Sign in",
        authenticated_marker_visible=True,
    )


def test_login_preflight_requires_authenticated_marker():
    assert not login_reached_authenticated_state(
        current_url="https://example.test/ITOAS/operacloud/mfa",
        password_visible=False,
        body_text="Enter verification code",
        authenticated_marker_visible=False,
    )
    assert login_reached_authenticated_state(
        current_url="https://example.test/ITOAS/operacloud/home",
        password_visible=False,
        body_text="Dashboard",
        authenticated_marker_visible=True,
    )


def test_read_reservation_workbook_key_value_shape(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path)

    records = read_reservation_records(workbook_path, "Rezervasyon 2", "Kodlar")

    assert len(records) == 1
    record = records[0]
    assert record.fields["arrival_date"] == "2027-06-02"
    assert record.fields["departure_date"] == "2027-06-04"
    assert record.fields["guest_name"] == "KOSEREISOGLU, EMIR MURAT"
    assert record.fields["children_count"] == 0
    assert validate_reservation_fields(record.fields, record.code_lists) == []


@pytest.mark.asyncio
async def test_opera_rezervasyon_workflow_writes_dry_run_report(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path)
    output_json = tmp_path / "runs" / "reservation_plan.json"
    output_html = tmp_path / "reports" / "reservation_plan.html"
    config = HarnessConfig(
        variables={
            "opera_input_excel": str(workbook_path),
            "opera_output_json": str(output_json),
            "opera_output_html": str(output_html),
        }
    )
    workflow = OperaRezervasyonFromExcelWorkflow(config=config)

    result = await workflow._execute()

    assert result.passed
    assert result.total_records == 1
    assert output_json.exists()
    assert output_html.exists()
    report = json.loads(output_json.read_text(encoding="utf-8"))
    assert report["dry_run"] is True
    assert report["valid_records"] == 1
    assert report["records"][0]["planned_stages"][0]["stage"] == "INIT"


@pytest.mark.asyncio
async def test_opera_rezervasyon_workflow_reports_missing_workbook(tmp_path):
    output_json = tmp_path / "runs" / "reservation_plan.json"
    output_html = tmp_path / "reports" / "reservation_plan.html"
    config = HarnessConfig(
        variables={
            "opera_input_excel": str(tmp_path / "missing.xlsx"),
            "opera_output_json": str(output_json),
            "opera_output_html": str(output_html),
        }
    )
    workflow = OperaRezervasyonFromExcelWorkflow(config=config)

    result = await workflow._execute()

    assert not result.passed
    assert "Workbook not found" in result.error_message
    assert output_json.exists()
    assert output_html.exists()


def test_opera_rezervasyon_validation_rejects_bad_code(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path, market="BAD")
    record = read_reservation_records(workbook_path, "Rezervasyon 2", "Kodlar")[0]

    errors = validate_reservation_fields(record.fields, record.code_lists)

    assert "market 'BAD' is not listed in market_codes" in errors


def test_opera_rezervasyon_validation_rejects_bad_rate_code(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path, rate_code="BADRATE")
    record = read_reservation_records(workbook_path, "Rezervasyon 2", "Kodlar")[0]

    errors = validate_reservation_fields(record.fields, record.code_lists)

    assert "rate_code 'BADRATE' is not listed in rate_codes" in errors


def test_opera_rezervasyon_validation_requires_positive_occupancy(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path, adults=0)
    record = read_reservation_records(workbook_path, "Rezervasyon 2", "Kodlar")[0]

    errors = validate_reservation_fields(record.fields, record.code_lists)

    assert "adults must be greater than zero" in errors


def test_opera_rezervasyon_validation_rejects_bad_rate(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path, rate="abc")
    record = read_reservation_records(workbook_path, "Rezervasyon 2", "Kodlar")[0]

    errors = validate_reservation_fields(record.fields, record.code_lists)

    assert "rate must be a positive number" in errors


def test_opera_rezervasyon_validation_rejects_bad_dates(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path, arrival="tomorrow")
    record = read_reservation_records(workbook_path, "Rezervasyon 2", "Kodlar")[0]

    errors = validate_reservation_fields(record.fields, record.code_lists)

    assert "arrival_date must be a valid ISO date" in errors


def test_opera_rezervasyon_validation_requires_code_lists(tmp_path):
    workbook_path = make_reservation_workbook(tmp_path, include_codes=False)
    record = read_reservation_records(workbook_path, "Rezervasyon 2", "Kodlar")[0]

    errors = validate_reservation_fields(record.fields, record.code_lists)

    assert "missing code list: market_codes" in errors
    assert "missing code list: room_types" in errors


def test_missing_credential_names_uses_secret_env_names_only():
    assert missing_credential_names(username=None, password="set") == ["OPERA_USERNAME"]
    assert missing_credential_names(username="set", password=None) == ["OPERA_PASSWORD"]


def make_reservation_workbook(
    tmp_path,
    *,
    market: str = "CIN",
    rate_code: str = "CORPBB",
    rate: str = "138.75",
    adults: int = 2,
    arrival=datetime(2027, 6, 2),
    include_codes: bool = True,
) -> Path:
    path = tmp_path / "reservation.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rezervasyon 2"
    rows = [
        ("Arrival", arrival, "ZORUNLU"),
        ("Departure", datetime(2027, 6, 4), "ZORUNLU"),
        ("Rooms", 1, None),
        ("Adults ", adults, None),
        ("Children", 0, None),
        ("Ages of Children", None, None),
        ("Name", "KOSEREISOGLU, EMIR MURAT", "ZORUNLU"),
        ("Company", "SECURITAS", None),
        ("Travel Agency", "PASSANGER", None),
        ("Block Code", None, None),
        ("Group", None, None),
        ("Room Feature", None, None),
        ("Room Type", "BDS", "ZORUNLU"),
        ("Room", None, None),
        ("Rate Codes", rate_code, "ZORUNLU"),
        ("Rate Category", None, None),
        ("Reservation Type", "SELFP", "ZORUNLU"),
        ("Market", market, "ZORUNLU"),
        ("Source", "COR", "ZORUNLU"),
        ("Room Type to Charge", "BDS", "ZORUNLU"),
        ("Rate", rate, "ZORUNLU"),
        ("Commision Amount", None, None),
        ("Commision %", None, None),
    ]
    for index, row in enumerate(rows, start=4):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=index, column=column, value=value)

    if include_codes:
        codes = workbook.create_sheet("Kodlar")
        codes.append(["Market Codes", "Source Codes", "Rate Codes", "Reservation Types", "Room Types"])
        codes.append(["CIN", "COR", "CORPBB", "SELFP", "BDS"])
    workbook.save(path)
    return path
