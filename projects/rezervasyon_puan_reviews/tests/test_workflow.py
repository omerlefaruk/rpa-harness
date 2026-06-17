from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).resolve().parents[1] / "workflow.py"
SPEC = importlib.util.spec_from_file_location("rezervasyon_puan_reviews", MODULE_PATH)
workflow = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
sys.modules[SPEC.name] = workflow
SPEC.loader.exec_module(workflow)


def test_read_rezervasyon_puan_records_keeps_score_count_and_link(tmp_path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Taksim Analiz"
    ws.cell(row=1, column=2, value="Expedia/Hotels")
    ws.cell(row=2, column=1, value="Otel Adı")
    ws.cell(row=2, column=2, value="Puan")
    ws.cell(row=2, column=3, value="Yorum")
    ws.cell(row=2, column=4, value="Link")
    ws.cell(row=3, column=1, value="The Marmara Taksim")
    ws.cell(row=3, column=2, value=9.0)
    ws.cell(row=3, column=3, value=1014)
    ws.cell(row=3, column=4, value="Git")
    ws.cell(row=3, column=4).hyperlink = "https://example.test/expedia"
    path = tmp_path / "Branches.xlsx"
    workbook.save(path)

    records = workflow.read_rezervasyon_puan_records(path, "Taksim Analiz")

    assert records == [
        {
            "id": "3:Expedia/Hotels",
            "source_row": 3,
            "hotel": "The Marmara Taksim",
            "platform": "Expedia/Hotels",
            "platform_score": 9.0,
            "platform_review_count": 1014,
            "url": "https://example.test/expedia",
            "domain": "example.test",
        }
    ]


def test_write_review_workbook_writes_summary_and_reviews(tmp_path):
    output = tmp_path / "reviews.xlsx"
    result = {
        "last_30_days_window": {"start": "2026-05-18", "end": "2026-06-17"},
        "summary": [
            {
                "hotel": "The Marmara Taksim",
                "platform": "Expedia/Hotels",
                "platform_score": 9.0,
                "platform_review_count": 1014,
                "status": "processed",
                "fetch_status": "loaded",
                "recent_review_count": 1,
                "blocked": False,
                "url": "https://example.test/expedia",
                "final_url": "https://example.test/expedia",
                "error": "",
            }
        ],
        "reviews": [
            {
                "hotel": "The Marmara Taksim",
                "platform": "Expedia/Hotels",
                "source_row": 3,
                "domain": "example.test",
                "date": "2026-06-10",
                "reviewer": "Guest",
                "rating": "10/10",
                "title": "Great stay",
                "text": "Clean rooms and helpful staff.",
                "source_url": "https://example.test/expedia",
                "extraction_method": "playwright_body_text",
            }
        ],
    }

    workflow.write_review_workbook(output, result)

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Summary", "Reviews"]
    assert workbook["Summary"]["A3"].value == "Hotel"
    assert workbook["Summary"]["D4"].value == 1014
    assert workbook["Reviews"]["A2"].value == "The Marmara Taksim"
    assert workbook["Reviews"]["H2"].value == "Clean rooms and helpful staff."


def test_canonical_url_removes_tracking_query():
    assert workflow.canonical_url("https://www.expedia.com/foo/bar?searchId=abc#x") == (
        "https://www.expedia.com/foo/bar"
    )


def test_full_plan_groups_expedia_cdp_then_headless():
    records = [
        {"platform": "Google", "domain": "google.com", "url": "https://google.com"},
        {"platform": "Expedia/Hotels", "domain": "www.expedia.com", "url": "https://expedia.com"},
    ]

    planned = workflow.build_record_plan(records, "full")

    assert [record["capture_method"] for record in planned] == [
        "cdp_chrome_profile",
        "headless_standard",
    ]


def test_dedupe_review_rows_keeps_unique_text():
    rows = [
        {"hotel": "A", "platform": "Google", "date": "2026-06-01", "reviewer": "R", "rating": "5/5", "text": "Nice"},
        {"hotel": "A", "platform": "Google", "date": "2026-06-01", "reviewer": "R", "rating": "5/5", "text": " Nice "},
        {"hotel": "A", "platform": "Google", "date": "2026-06-02", "reviewer": "R", "rating": "5/5", "text": "Nice"},
    ]

    assert len(workflow.dedupe_review_rows(rows)) == 2


def test_block_reason_classifies_known_failures():
    assert workflow.block_reason({"title": "Bot or Not?", "error": "", "fetch_status": "blocked_or_empty"}) == (
        "Expedia anti-bot page: Bot or Not"
    )
    assert workflow.block_reason({"title": "", "error": "HTTP Error 403: Forbidden", "fetch_status": "failed"}) == (
        "HTTP 403 Forbidden"
    )


def test_write_raw_outputs_writes_method_and_combined_files(tmp_path):
    result = {
        "input_excel": "input.xlsx",
        "sheet": "Taksim Analiz",
        "run_id": "20260617_000000",
        "run_mode": "full",
        "last_30_days_window": {"start": "2026-05-18", "end": "2026-06-17"},
        "summary": [
            {"hotel": "A", "platform": "Expedia/Hotels", "capture_method": "cdp_chrome_profile"},
            {"hotel": "A", "platform": "Google", "capture_method": "headless_standard"},
        ],
        "reviews": [
            {"hotel": "A", "platform": "Expedia/Hotels", "capture_method": "cdp_chrome_profile"},
            {"hotel": "A", "platform": "Google", "capture_method": "headless_standard"},
        ],
    }

    files = workflow.write_raw_outputs(tmp_path, "20260617_000000", result)

    assert {path.name for path in files} == {
        "20260617_000000_cdp_chrome_profile.json",
        "20260617_000000_headless_standard.json",
        "20260617_000000_combined.json",
    }
