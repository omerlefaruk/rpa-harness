from __future__ import annotations

import asyncio
import importlib.util
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse, urlunparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = "C:/Users/Rau/Desktop/Archive/TM-57-Rezervasyon_Puan/data/Branches.xlsx"
DEFAULT_SHEET = "Taksim Analiz"
DEFAULT_JSON = "runs/rezervasyon_puan_reviews/reviews_last_30_days.json"
DEFAULT_HTML = "reports/rezervasyon_puan_reviews/reviews_last_30_days.html"
DEFAULT_XLSX = "reports/rezervasyon_puan_reviews/reviews_last_30_days.xlsx"
PLATFORM_START_COLUMN = 2
PLATFORM_WIDTH = 3
HOTEL_NAME_COLUMN = 1


def _load_legacy_ota_module():
    path = Path(__file__).resolve().parent / "_review_collector.py"
    spec = importlib.util.spec_from_file_location("rezervasyon_review_collector", path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


legacy_ota = _load_legacy_ota_module()


class RezervasyonPuanReviewsWorkflow(legacy_ota.OtaRecentReviewsFromExcelWorkflow):
    name = "rezervasyon_puan_reviews"
    tags = ["rpa", "excel", "browser", "reviews", "ota", "tm-57"]
    max_retries_per_record = 0

    async def setup(self):
        variables = getattr(self.config, "variables", {}) or {}
        self.input_path = Path(variables.get("input_excel", DEFAULT_INPUT))
        self.sheet_name = variables.get("sheet", DEFAULT_SHEET)
        self.as_of = legacy_ota._as_of_date(variables.get("as_of_date"))
        self.start_date = self.as_of - timedelta(days=30)
        self.output_json = Path(variables.get("output_json", DEFAULT_JSON))
        self.output_html = Path(variables.get("output_html", DEFAULT_HTML))
        self.output_xlsx = Path(variables.get("output_excel", DEFAULT_XLSX))
        self.raw_output_dir = Path(variables.get("raw_output_dir", "runs/rezervasyon_puan_reviews/raw"))
        self.evidence_dir = Path(variables.get("evidence_dir", "runs/rezervasyon_puan_reviews/evidence"))
        self.run_id = variables.get("run_id") or datetime.now().strftime("%Y%m%d_%H%M%S")
        self.run_mode = str(variables.get("run_mode", "full")).lower()
        self.timeout_ms = int(variables.get("timeout_ms", legacy_ota.PAGE_TIMEOUT_MS))
        self.browser_profile_dir = variables.get("browser_profile_dir")
        self.default_browser_profile_dir = self.browser_profile_dir
        self.browser_headed = as_bool(variables.get("browser_headed", False))
        self.default_browser_headed = self.browser_headed
        self.browser_cdp = as_bool(variables.get("browser_cdp", False))
        self.default_browser_cdp = self.browser_cdp
        self.browser_cdp_port = int(variables.get("browser_cdp_port", 9223))
        self.expedia_canonical_fallback = as_bool(variables.get("expedia_canonical_fallback", True))
        self.google_scrolls = int(variables.get("google_scrolls", 12))
        self.debug_evidence = self.run_mode == "debug" or as_bool(variables.get("debug_evidence", False))
        self.records = build_record_plan(
            read_rezervasyon_puan_records(self.input_path, self.sheet_name),
            self.run_mode,
        )
        self.summary_rows: list[dict] = []
        self.review_rows: list[dict] = []
        self.playwright = None
        self.browser = None
        self.page = None
        self.browser_process = None
        self.active_capture_method = None
        self.log(
            f"Loaded {len(self.records)} Rezervasyon Puan links from "
            f"{self.input_path} / {self.sheet_name}; review window "
            f"{self.start_date.isoformat()} to {self.as_of.isoformat()}; mode={self.run_mode}"
        )

    async def process_record(self, record: dict) -> dict:
        await self.configure_capture_method(record.get("capture_method", capture_method_for(record)))
        before = len(self.review_rows)
        result = await super().process_record(record)
        summary = self.summary_rows[-1]
        summary["capture_method"] = record.get("capture_method", "")
        summary["block_reason"] = block_reason(summary)
        for review in self.review_rows[before:]:
            review["capture_method"] = summary["capture_method"]
            if is_expedia(record):
                review["extraction_method"] = "expedia_page_reviews"
        if should_capture_evidence(summary, self.debug_evidence):
            await self.capture_evidence(summary)
        return result

    async def configure_capture_method(self, method: str) -> None:
        if method == self.active_capture_method:
            return
        await self.close_browser()
        self.active_capture_method = method
        if method == "cdp_chrome_profile":
            self.browser_cdp = True
            self.browser_headed = True
            self.browser_profile_dir = self.default_browser_profile_dir or (
                "runs/rezervasyon_puan_reviews/chrome_cdp_workflow_profile"
            )
        else:
            self.browser_cdp = False
            self.browser_headed = False
            self.browser_profile_dir = None

    async def close_browser(self) -> None:
        try:
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()
        finally:
            if self.browser_process:
                self.browser_process.terminate()
                try:
                    self.browser_process.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    self.browser_process.kill()
            self.playwright = None
            self.browser = None
            self.page = None
            self.browser_process = None

    async def ensure_browser(self) -> None:
        if self.page:
            return
        from playwright.async_api import async_playwright

        self.playwright = await async_playwright().start()
        if self.browser_cdp:
            chrome = chrome_exe()
            profile = Path(self.browser_profile_dir or "runs/rezervasyon_puan_reviews/chrome_cdp_profile").resolve()
            profile.mkdir(parents=True, exist_ok=True)
            active_port = profile / "DevToolsActivePort"
            if active_port.exists():
                active_port.unlink()
            self.browser_process = subprocess.Popen(
                [
                    str(chrome),
                    f"--remote-debugging-port={self.browser_cdp_port}",
                    f"--user-data-dir={profile}",
                    "--no-first-run",
                    "--disable-default-apps",
                    "about:blank",
                ]
            )
            self.browser = await self.connect_over_cdp(profile)
            context = self.browser.contexts[0]
            self.page = context.pages[0] if context.pages else await context.new_page()
            return

        kwargs = {
            "headless": not self.browser_headed,
            "locale": "tr-TR",
            "user_agent": legacy_ota.trip_reviews.public_request_headers()["User-Agent"],
        }
        if self.browser_profile_dir:
            Path(self.browser_profile_dir).mkdir(parents=True, exist_ok=True)
            try:
                self.browser = await self.playwright.chromium.launch_persistent_context(
                    str(self.browser_profile_dir),
                    channel="chrome",
                    **kwargs,
                )
            except Exception:
                self.browser = await self.playwright.chromium.launch_persistent_context(
                    str(self.browser_profile_dir),
                    **kwargs,
                )
            self.page = self.browser.pages[0] if self.browser.pages else await self.browser.new_page()
            return

        self.browser = await self.playwright.chromium.launch(headless=not self.browser_headed)
        context = await self.browser.new_context(
            locale=kwargs["locale"],
            user_agent=kwargs["user_agent"],
        )
        self.page = await context.new_page()

    async def connect_over_cdp(self, profile: Path):
        last_error = None
        for _ in range(30):
            port = self.browser_cdp_port
            if port == 0:
                active_port = profile / "DevToolsActivePort"
                if active_port.exists():
                    port = int(active_port.read_text(encoding="utf-8").splitlines()[0])
            if port:
                try:
                    return await self.playwright.chromium.connect_over_cdp(
                        f"http://127.0.0.1:{port}"
                    )
                except Exception as exc:
                    last_error = exc
            await asyncio.sleep(0.5)
        if last_error:
            raise last_error
        raise RuntimeError("Chrome did not expose a CDP port")

    async def fetch_text(self, record: dict) -> tuple[str, dict]:
        if is_tripadvisor(record):
            text, status = legacy_ota.trip_reviews.fetch_public_text_with_status(record["url"])
            if text and not legacy_ota.looks_blocked(text):
                return text, {
                    "status": "loaded",
                    "fetch_method": "http_public_html",
                    "final_url": record["url"],
                    "title": "",
                    "text_length": len(text),
                    "blocked": False,
                    "error": "",
                }
            return "", {
                "status": "failed",
                "fetch_method": "http_public_html",
                "final_url": record["url"],
                "title": "",
                "text_length": 0,
                "blocked": False,
                "error": status.get("error") or status.get("status") or "Tripadvisor fetch failed",
            }

        text, status = await super().fetch_text(record)
        if not (
            self.expedia_canonical_fallback
            and is_expedia(record)
            and status.get("blocked")
        ):
            return text, status

        fallback_url = canonical_url(record["url"])
        if fallback_url == record["url"]:
            return text, status

        fallback_text, fallback_status = await self.fetch_text_with_browser(fallback_url)
        fallback_status["fetch_method"] = f"{fallback_status.get('fetch_method', '')}+canonical_url"
        if fallback_text and not fallback_status.get("blocked"):
            return fallback_text, fallback_status

        status["fallback_url"] = fallback_url
        status["fallback_status"] = fallback_status.get("status")
        status["fallback_title"] = fallback_status.get("title")
        return text, status

    async def scroll_google_reviews(self) -> None:
        for _ in range(self.google_scrolls):
            try:
                await self.page.mouse.wheel(0, 1800)
            except Exception:
                await self.page.evaluate("() => window.scrollBy(0, 1800)")
            await self.page.wait_for_timeout(700)

    async def capture_evidence(self, summary: dict) -> None:
        self.evidence_dir.mkdir(parents=True, exist_ok=True)
        base = self.evidence_dir / safe_name(
            f"{self.run_id}_{summary.get('source_row')}_{summary.get('hotel')}_{summary.get('platform')}"
        )
        text = ""
        if self.page:
            try:
                text = await self.page.locator("body").inner_text(timeout=3000)
                screenshot = base.with_suffix(".png")
                await self.page.screenshot(path=str(screenshot), full_page=True)
                summary["evidence_screenshot"] = str(screenshot)
            except Exception as exc:
                summary["evidence_error"] = legacy_ota.sanitize_error(exc)
        evidence = {
            "hotel": summary.get("hotel"),
            "platform": summary.get("platform"),
            "status": summary.get("status"),
            "fetch_status": summary.get("fetch_status"),
            "block_reason": summary.get("block_reason"),
            "error": summary.get("error"),
            "final_url": summary.get("final_url"),
            "title": summary.get("title"),
            "text_preview": legacy_ota.normalize_text(text)[:1200],
        }
        evidence_path = base.with_suffix(".json")
        evidence_path.write_text(json.dumps(evidence, indent=2, ensure_ascii=False), encoding="utf-8")
        summary["evidence_json"] = str(evidence_path)

    async def teardown(self):
        await self.close_browser()
        self.review_rows = dedupe_review_rows(self.review_rows)

        result = {
            "input_excel": str(self.input_path),
            "sheet": self.sheet_name,
            "run_id": self.run_id,
            "run_mode": self.run_mode,
            "last_30_days_window": {
                "start": self.start_date.isoformat(),
                "end": self.as_of.isoformat(),
            },
            "total_links": len(self.summary_rows),
            "processed": sum(1 for row in self.summary_rows if row["status"] == "processed"),
            "errors": sum(1 for row in self.summary_rows if row["status"] == "error"),
            "blocked_or_empty": sum(1 for row in self.summary_rows if row.get("blocked")),
            "failed_or_blocked": sum(
                1
                for row in self.summary_rows
                if row.get("blocked") or row.get("fetch_status") in {"failed", "blocked_or_empty"}
            ),
            "links_with_recent_reviews": sum(1 for row in self.summary_rows if row["recent_review_count"] > 0),
            "total_recent_reviews": len(self.review_rows),
            "summary": self.summary_rows,
            "reviews": self.review_rows,
        }
        self.output_json.parent.mkdir(parents=True, exist_ok=True)
        self.output_html.parent.mkdir(parents=True, exist_ok=True)
        self.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        self.raw_output_dir.mkdir(parents=True, exist_ok=True)
        raw_files = write_raw_outputs(self.raw_output_dir, self.run_id, result)
        result["raw_output_files"] = [str(path) for path in raw_files]
        self.output_json.write_text(
            json.dumps(result, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        self.output_html.write_text(render_html_report(result), encoding="utf-8")
        write_review_workbook(self.output_xlsx, result)
        self.result.output_files.extend(
            [str(self.output_json), str(self.output_html), str(self.output_xlsx), *result["raw_output_files"]]
        )
        self.log(f"Wrote JSON: {self.output_json}")
        self.log(f"Wrote HTML: {self.output_html}")
        self.log(f"Wrote XLSX: {self.output_xlsx}")


def read_rezervasyon_puan_records(input_path: str | Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(input_path, data_only=False)
    worksheet = workbook[sheet_name]
    records: list[dict] = []
    for row in range(3, worksheet.max_row + 1):
        hotel = worksheet.cell(row=row, column=HOTEL_NAME_COLUMN).value
        if not hotel:
            continue
        for platform, score_column, count_column, link_column in platform_columns(worksheet):
            link_cell = worksheet.cell(row=row, column=link_column)
            target = link_cell.hyperlink.target if link_cell.hyperlink else None
            if not target:
                continue
            records.append(
                {
                    "id": f"{row}:{platform}",
                    "source_row": row,
                    "hotel": str(hotel).strip(),
                    "platform": platform,
                    "platform_score": as_float(worksheet.cell(row=row, column=score_column).value),
                    "platform_review_count": as_int(worksheet.cell(row=row, column=count_column).value),
                    "url": target,
                    "domain": urlparse(target).netloc,
                }
            )
    return records


def platform_columns(worksheet) -> list[tuple[str, int, int, int]]:
    columns: list[tuple[str, int, int, int]] = []
    column = PLATFORM_START_COLUMN
    while column <= worksheet.max_column:
        platform = worksheet.cell(row=1, column=column).value
        if platform and str(platform).strip().lower() != "ortalama":
            columns.append((str(platform).strip(), column, column + 1, column + 2))
        column += PLATFORM_WIDTH
    return columns


def build_record_plan(records: list[dict], run_mode: str) -> list[dict]:
    planned = [{**record, "capture_method": capture_method_for(record)} for record in records]
    if run_mode in {"full", "debug"}:
        return [
            *[record for record in planned if record["capture_method"] == "cdp_chrome_profile"],
            *[record for record in planned if record["capture_method"] != "cdp_chrome_profile"],
        ]
    if run_mode == "headless":
        return [{**record, "capture_method": "headless_standard"} for record in records]
    if run_mode == "cdp":
        return [{**record, "capture_method": "cdp_chrome_profile"} for record in records]
    return planned


def capture_method_for(record: dict) -> str:
    return "cdp_chrome_profile" if is_expedia(record) else "headless_standard"


def is_expedia(record: dict) -> bool:
    return "expedia." in record.get("domain", "") or "hotels." in record.get("domain", "")


def is_tripadvisor(record: dict) -> bool:
    platform = str(record.get("platform", "")).lower()
    return "tripadvisor" in platform or "tripadvisor." in record.get("domain", "")


def canonical_url(url: str) -> str:
    parsed = urlparse(url)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", "", ""))


def chrome_exe() -> Path:
    candidates = [
        Path(os.environ.get("PROGRAMFILES", "")) / "Google/Chrome/Application/chrome.exe",
        Path(os.environ.get("PROGRAMFILES(X86)", "")) / "Google/Chrome/Application/chrome.exe",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("Google Chrome executable was not found")


def should_capture_evidence(summary: dict, debug_evidence: bool) -> bool:
    return debug_evidence or bool(summary.get("blocked")) or summary.get("fetch_status") in {"failed", "blocked_or_empty"}


def block_reason(summary: dict) -> str:
    text = " ".join(
        str(summary.get(key) or "")
        for key in ("title", "error", "fetch_status", "final_url")
    ).lower()
    if "bot or not" in text or "human side" in text:
        return "Expedia anti-bot page: Bot or Not"
    if "403" in text or "forbidden" in text:
        return "HTTP 403 Forbidden"
    if summary.get("blocked"):
        return "Blocked by site anti-bot or verification page"
    if summary.get("fetch_status") == "blocked_or_empty":
        return "Blocked or empty page"
    if summary.get("fetch_status") == "failed":
        return str(summary.get("error") or "Fetch failed")
    return ""


def safe_name(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", value).strip("_")[:140] or "evidence"


def dedupe_review_rows(rows: list[dict]) -> list[dict]:
    seen: set[tuple[str, str, str, str, str, str]] = set()
    unique: list[dict] = []
    for row in rows:
        key = tuple(
            re.sub(r"\s+", " ", str(row.get(field) or "")).strip().lower()
            for field in ("hotel", "platform", "date", "reviewer", "rating", "text")
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def write_raw_outputs(output_dir: Path, run_id: str, result: dict) -> list[Path]:
    files: list[Path] = []
    for method in sorted({row.get("capture_method", "unknown") for row in result["summary"]}):
        payload = {
            **{key: result[key] for key in ("input_excel", "sheet", "run_id", "run_mode", "last_30_days_window")},
            "capture_method": method,
            "summary": [row for row in result["summary"] if row.get("capture_method") == method],
            "reviews": [row for row in result["reviews"] if row.get("capture_method") == method],
        }
        path = output_dir / f"{run_id}_{safe_name(method)}.json"
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        files.append(path)
    combined = output_dir / f"{run_id}_combined.json"
    combined.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    files.append(combined)
    return files


def render_html_report(result: dict) -> str:
    blocked_rows = "".join(
        "<tr>"
        f"<td>{row.get('hotel', '')}</td>"
        f"<td>{row.get('platform', '')}</td>"
        f"<td>{row.get('fetch_status', '')}</td>"
        f"<td>{row.get('block_reason', '')}</td>"
        f"<td>{row.get('evidence_json', '')}</td>"
        "</tr>"
        for row in result["summary"]
        if row.get("block_reason")
    )
    return (
        "<!doctype html><html><head><meta charset=\"utf-8\">"
        "<title>Rezervasyon Puan Reviews</title></head><body>"
        "<h1>Rezervasyon Puan Reviews</h1>"
        f"<p>{result['last_30_days_window']['start']} to {result['last_30_days_window']['end']} "
        f"| mode: {result.get('run_mode', '')} | run: {result.get('run_id', '')}</p>"
        f"<p>Total links: {result['total_links']} | Recent reviews: {result['total_recent_reviews']} "
        f"| Failed/blocked: {result.get('failed_or_blocked', 0)}</p>"
        "<h2>Blocked / Failed</h2>"
        "<table border=\"1\" cellspacing=\"0\" cellpadding=\"4\"><thead><tr>"
        "<th>Hotel</th><th>Platform</th><th>Status</th><th>Reason</th><th>Evidence</th>"
        f"</tr></thead><tbody>{blocked_rows}</tbody></table>"
        "</body></html>"
    )


def write_review_workbook(output_path: Path, result: dict) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(
        [
            "Window Start",
            result["last_30_days_window"]["start"],
            "Window End",
            result["last_30_days_window"]["end"],
            "Run Mode",
            result.get("run_mode", ""),
            "Run ID",
            result.get("run_id", ""),
        ]
    )
    summary.append([])
    summary.append(
        [
            "Hotel",
            "Platform",
            "Platform Score",
            "Platform Review Count",
            "Recent Review Count",
            "Status",
            "Fetch Status",
            "Capture Method",
            "Blocked",
            "Block Reason",
            "URL",
            "Final URL",
            "Error",
            "Evidence JSON",
            "Evidence Screenshot",
        ]
    )
    for row in result["summary"]:
        summary.append(
            [
                row.get("hotel"),
                row.get("platform"),
                row.get("platform_score"),
                row.get("platform_review_count"),
                row.get("recent_review_count"),
                row.get("status"),
                row.get("fetch_status"),
                row.get("capture_method"),
                row.get("blocked"),
                row.get("block_reason"),
                row.get("url"),
                row.get("final_url"),
                row.get("error"),
                row.get("evidence_json"),
                row.get("evidence_screenshot"),
            ]
        )

    reviews = workbook.create_sheet("Reviews")
    reviews.append(
        [
            "Hotel",
            "Platform",
            "Source Row",
            "Domain",
            "Review Date",
            "Reviewer",
            "Rating",
            "Review Text",
            "Title",
            "Source URL",
            "Extraction Method",
            "Capture Method",
        ]
    )
    for row in result["reviews"]:
        reviews.append(
            [
                row.get("hotel"),
                row.get("platform"),
                row.get("source_row"),
                row.get("domain"),
                row.get("date"),
                row.get("reviewer"),
                row.get("rating"),
                row.get("text"),
                row.get("title"),
                row.get("source_url"),
                row.get("extraction_method"),
                row.get("capture_method"),
            ]
        )

    for worksheet in workbook.worksheets:
        header_row = 3 if worksheet.title == "Summary" else 1
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True)
        for column_cells in worksheet.columns:
            max_len = min(max(len(str(cell.value or "")) for cell in column_cells), 80)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, max_len + 2)
        worksheet.freeze_panes = "A4" if worksheet.title == "Summary" else "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)


def as_float(value) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def as_int(value) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def as_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}
