"""Unit tests for the OTA link swarm Excel workflow helpers."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook


MODULE_PATH = Path(__file__).parent / "rpa" / "ota_link_swarm_from_excel.py"
SPEC = importlib.util.spec_from_file_location("ota_link_swarm_from_excel", MODULE_PATH)
ota_workflow = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
sys.modules[SPEC.name] = ota_workflow
SPEC.loader.exec_module(ota_workflow)


def test_read_ota_link_records_reads_platform_hyperlinks(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Taksim Analiz"
    worksheet.cell(row=1, column=2, value="Expedia/Hotels")
    worksheet.cell(row=1, column=5, value="Trip.com")
    worksheet.cell(row=1, column=8, value="Ortalama")
    worksheet.cell(row=3, column=1, value="The Marmara Taksim")
    worksheet.cell(row=3, column=4, value="Expedia")
    worksheet.cell(row=3, column=4).hyperlink = "https://example.com/expedia"
    worksheet.cell(row=3, column=7, value="Trip")
    worksheet.cell(row=3, column=7).hyperlink = "https://example.com/trip"
    worksheet.cell(row=3, column=10, value="Average")
    worksheet.cell(row=3, column=10).hyperlink = "https://example.com/ignored"
    path = tmp_path / "ota.xlsx"
    workbook.save(path)

    records = ota_workflow.read_ota_link_records(path, "Taksim Analiz")

    assert [(record["hotel"], record["platform"], record["url"]) for record in records] == [
        ("The Marmara Taksim", "Expedia/Hotels", "https://example.com/expedia"),
        ("The Marmara Taksim", "Trip.com", "https://example.com/trip"),
    ]


def test_relative_or_raw_links_artifacts_from_report_folder(tmp_path):
    report_path = tmp_path / "reports" / "ota_link_swarm_report.html"
    artifact_path = tmp_path / "runs" / "ota_link_swarm" / "link" / "report.html"
    report_path.parent.mkdir()
    artifact_path.parent.mkdir(parents=True)
    artifact_path.write_text("artifact", encoding="utf-8")

    assert ota_workflow.relative_or_raw(str(artifact_path), result_path=str(report_path)) == (
        "../runs/ota_link_swarm/link/report.html"
    )
