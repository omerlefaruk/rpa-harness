"""
Excel-driven OPERA Cloud reservation workflow scaffold.

This workflow reads the provided reservation workbook, validates the booking
fields and code lists, then writes a deterministic dry-run plan. It does not
submit reservations; live OPERA browser actions should be enabled only after
selectors and a safe test account are verified.
"""

from __future__ import annotations

import html
import json
import os
import re
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from harness import RPAWorkflow
from harness.security import redact_value


DEFAULT_INPUT = "data/operaRezervasyon/Rezervasyon Excel - Robot.xlsx"
DEFAULT_SHEET = "Rezervasyon 2"
DEFAULT_CODES_SHEET = "Kodlar"
DEFAULT_OUTPUT_JSON = "runs/operaRezervasyon/reservation_plan.json"
DEFAULT_OUTPUT_HTML = "reports/operaRezervasyon/reservation_plan.html"
DEFAULT_EVIDENCE_DIR = "runs/operaRezervasyon/evidence"

FIELD_ALIASES = {
    "arrival": "arrival_date",
    "departure": "departure_date",
    "rooms": "rooms",
    "adults": "adults",
    "adults ": "adults",
    "children": "children_count",
    "ages of children": "children_ages",
    "name": "guest_name",
    "company": "company",
    "travel agency": "travel_agency",
    "block code": "block_code",
    "group": "group",
    "room feature": "room_feature",
    "room type": "room_type",
    "room": "room",
    "rate codes": "rate_code",
    "rate category": "rate_category",
    "reservation type": "reservation_type",
    "market": "market",
    "source": "source",
    "room type to charge": "room_type_to_charge",
    "rate": "rate",
    "commision amount": "commission_amount",
    "commission amount": "commission_amount",
    "commision %": "commission_percent",
    "commission %": "commission_percent",
}

REQUIRED_FIELDS = [
    "arrival_date",
    "departure_date",
    "rooms",
    "adults",
    "guest_name",
    "room_type",
    "rate_code",
    "reservation_type",
    "market",
    "source",
    "room_type_to_charge",
    "rate",
]

CODE_FIELD_TO_LIST = {
    "market": "market_codes",
    "source": "source_codes",
    "rate_code": "rate_codes",
    "reservation_type": "reservation_types",
    "room_type": "room_types",
    "room_type_to_charge": "room_types",
}

STAGES = [
    {
        "stage": "INIT",
        "intent": "Load workbook values and normalize reservation variables.",
        "success_checks": ["input workbook exists", "required sheet exists", "required fields present"],
    },
    {
        "stage": "LOGIN",
        "intent": "Open OPERA Cloud and authenticate with configured secret names.",
        "success_checks": ["URL contains /operacloud", "authenticated landing page visible"],
    },
    {
        "stage": "BOOKINGS",
        "intent": "Open the booking creation flow and enter stay dates, rooms, adults, and child count.",
        "success_checks": ["booking form visible", "arrival and departure fields match workbook"],
    },
    {
        "stage": "TOP_MENU_FILLINGS",
        "intent": "Fill room, rate, reservation type, market, source, charge room type, and rate values.",
        "success_checks": ["all OPERA code fields match workbook", "rate field matches workbook"],
    },
    {
        "stage": "CHILDREN",
        "intent": "Fill child age fields only when the workbook has child guests.",
        "success_checks": ["child age field count equals child count"],
    },
    {
        "stage": "GUEST_PROFILE",
        "intent": "Find or create the guest profile using the workbook guest name and optional company/agency.",
        "success_checks": ["guest profile selected", "guest name matches workbook"],
    },
    {
        "stage": "REST_OF_CODES",
        "intent": "Apply optional block, group, feature, room, commission, and category values.",
        "success_checks": ["optional populated values are visible in OPERA"],
    },
    {
        "stage": "REPORTING",
        "intent": "Capture final reservation reference and write run evidence.",
        "success_checks": ["confirmation reference captured", "report artifact exists"],
    },
]


@dataclass
class ReservationRecord:
    id: str
    fields: dict[str, Any]
    required_markers: dict[str, str] = field(default_factory=dict)
    source_rows: dict[str, int] = field(default_factory=dict)
    code_lists: dict[str, set[str]] = field(default_factory=dict)


class OperaRezervasyonFromExcelWorkflow(RPAWorkflow):
    name = "opera_rezervasyon_from_excel"
    tags = ["rpa", "excel", "opera", "reservation"]
    max_retries_per_record = 0

    async def setup(self):
        variables = getattr(self.config, "variables", {}) or {}
        self.input_path = Path(
            variables.get("opera_input_excel")
            or variables.get("input_excel")
            or os.getenv("OPERA_REZERVASYON_INPUT_EXCEL", DEFAULT_INPUT)
        )
        self.sheet_name = variables.get("opera_sheet", DEFAULT_SHEET)
        self.codes_sheet = variables.get("opera_codes_sheet", DEFAULT_CODES_SHEET)
        self.output_json = Path(variables.get("opera_output_json", DEFAULT_OUTPUT_JSON))
        self.output_html = Path(variables.get("opera_output_html", DEFAULT_OUTPUT_HTML))
        self.evidence_dir = Path(variables.get("opera_evidence_dir", DEFAULT_EVIDENCE_DIR))
        self.template_workflow = variables.get("opera_template_workflow", "")
        self.opera_url = variables.get("opera_url", "")
        self.live_preflight = bool_from_config(variables.get("opera_live_preflight", False))
        self.allow_login = bool_from_config(variables.get("opera_allow_login", False))
        self.authenticated_marker = variables.get("opera_authenticated_marker", "")
        self.rows: list[dict[str, Any]] = []
        self.records = read_reservation_records(self.input_path, self.sheet_name, self.codes_sheet)
        self.log(f"Loaded {len(self.records)} reservation record(s) from {self.input_path} / {self.sheet_name}")

    def get_records(self):
        for record in self.records:
            yield {
                "id": record.id,
                "fields": record.fields,
                "required_markers": record.required_markers,
                "source_rows": record.source_rows,
                "code_lists": record.code_lists,
            }

    async def process_record(self, record: dict) -> dict:
        fields = record["fields"]
        validation_errors = validate_reservation_fields(fields, record.get("code_lists", {}))
        plan = build_execution_plan(fields)
        live_evidence = {}
        if self.live_preflight and not validation_errors:
            live_evidence = await run_opera_live_preflight(
                opera_url=self.opera_url,
                evidence_dir=self.evidence_dir,
                allow_login=self.allow_login,
                authenticated_marker=self.authenticated_marker,
                headless=bool(getattr(self.config, "headless", True)),
                browser_name=str(getattr(self.config, "browser", "chromium")),
            )
            if live_evidence.get("status") != "passed":
                validation_errors.append(str(live_evidence.get("blocker") or "live preflight failed"))

        summary = {
            "record_id": record["id"],
            "status": "failed" if validation_errors else "passed",
            "guest_name": fields.get("guest_name", ""),
            "arrival_date": fields.get("arrival_date", ""),
            "departure_date": fields.get("departure_date", ""),
            "room_type": fields.get("room_type", ""),
            "rate_code": fields.get("rate_code", ""),
            "reservation_type": fields.get("reservation_type", ""),
            "market": fields.get("market", ""),
            "source": fields.get("source", ""),
            "children_count": fields.get("children_count", 0),
            "validation_errors": validation_errors,
            "live_preflight": live_evidence,
            "planned_stages": plan,
        }
        self.rows.append(summary)
        self.record_evidence(
            {
                "template_workflow": self.template_workflow,
                "opera_url_configured": bool(self.opera_url),
                "dry_run": not self.allow_login,
                "live_preflight": bool(live_evidence),
                "validation_errors": validation_errors,
                "planned_stage_count": len(plan),
            },
            record=record,
            stage="validate_reservation_record",
        )
        if validation_errors:
            return {
                "status": "failed",
                "reason": "; ".join(validation_errors),
                "details": redact_value(summary),
            }
        return {
            "status": "passed",
            "details": redact_value(
                {
                    "record_id": record["id"],
                    "guest_name": fields.get("guest_name"),
                    "arrival_date": fields.get("arrival_date"),
                    "departure_date": fields.get("departure_date"),
                    "planned_stage_count": len(plan),
                }
            ),
        }

    async def teardown(self):
        result = {
            "workflow": self.name,
            "input_excel": str(self.input_path),
            "sheet": self.sheet_name,
            "template_workflow": self.template_workflow,
            "dry_run": not self.allow_login,
            "live_preflight_enabled": self.live_preflight,
            "allow_login": self.allow_login,
            "total_records": len(self.rows),
            "valid_records": sum(1 for row in self.rows if row["status"] == "passed"),
            "invalid_records": sum(1 for row in self.rows if row["status"] != "passed"),
            "records": self.rows,
        }
        self.output_json.parent.mkdir(parents=True, exist_ok=True)
        self.output_html.parent.mkdir(parents=True, exist_ok=True)
        self.output_json.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
        self.output_html.write_text(render_html_report(result), encoding="utf-8")
        self.result.output_files.extend([str(self.output_json), str(self.output_html)])
        self.log(f"Wrote JSON: {self.output_json}")
        self.log(f"Wrote HTML: {self.output_html}")


def read_reservation_records(input_path: Path, sheet_name: str, codes_sheet: str) -> list[ReservationRecord]:
    from openpyxl import load_workbook

    if not input_path.exists():
        raise FileNotFoundError(f"Workbook not found: {input_path}")

    workbook = load_workbook(input_path, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        code_lists = read_code_lists(workbook, codes_sheet)
        fields: dict[str, Any] = {}
        required_markers: dict[str, str] = {}
        source_rows: dict[str, int] = {}
        for row_number in range(1, worksheet.max_row + 1):
            raw_label = worksheet.cell(row=row_number, column=1).value
            if raw_label is None:
                continue
            field_name = FIELD_ALIASES.get(normalize_label(raw_label))
            if not field_name:
                continue
            value = normalize_field_value(field_name, worksheet.cell(row=row_number, column=2).value)
            fields[field_name] = value
            source_rows[field_name] = row_number
            marker = worksheet.cell(row=row_number, column=3).value
            if marker:
                required_markers[field_name] = str(marker)

        record_id = reservation_record_id(fields)
        return [
            ReservationRecord(
                id=record_id,
                fields=fields,
                required_markers=required_markers,
                source_rows=source_rows,
                code_lists=code_lists,
            )
        ]
    finally:
        workbook.close()


def read_code_lists(workbook: Any, codes_sheet: str) -> dict[str, set[str]]:
    if codes_sheet not in workbook.sheetnames:
        return {}
    worksheet = workbook[codes_sheet]
    headers = [normalize_code_header(cell.value) for cell in worksheet[1]]
    codes: dict[str, set[str]] = {header: set() for header in headers if header}
    for row in range(2, worksheet.max_row + 1):
        for index, header in enumerate(headers, start=1):
            value = worksheet.cell(row=row, column=index).value
            if header and value not in (None, ""):
                codes[header].add(str(value).strip().upper())
    return codes


def validate_reservation_fields(fields: dict[str, Any], code_lists: dict[str, set[str]]) -> list[str]:
    errors: list[str] = []
    for field_name in REQUIRED_FIELDS:
        if fields.get(field_name) in (None, ""):
            errors.append(f"missing required field: {field_name}")

    arrival = parse_iso_date(fields.get("arrival_date"))
    departure = parse_iso_date(fields.get("departure_date"))
    if fields.get("arrival_date") and not arrival:
        errors.append("arrival_date must be a valid ISO date")
    if fields.get("departure_date") and not departure:
        errors.append("departure_date must be a valid ISO date")
    if arrival and departure and departure <= arrival:
        errors.append("departure_date must be after arrival_date")

    children_count = int_or_zero(fields.get("children_count"))
    children_ages = fields.get("children_ages") or []
    for field_name in ("rooms", "adults"):
        if int_or_zero(fields.get(field_name)) <= 0:
            errors.append(f"{field_name} must be greater than zero")
    if not positive_decimal(fields.get("rate")):
        errors.append("rate must be a positive number")
    if children_count < 0:
        errors.append("children_count cannot be negative")
    if children_count and len(children_ages) != children_count:
        errors.append("children_ages count must match children_count")

    for code_list_name in sorted(set(CODE_FIELD_TO_LIST.values())):
        if not code_lists.get(code_list_name):
            errors.append(f"missing code list: {code_list_name}")

    for field_name, code_list_name in CODE_FIELD_TO_LIST.items():
        value = fields.get(field_name)
        allowed = code_lists.get(code_list_name)
        if value and allowed and str(value).strip().upper() not in allowed:
            errors.append(f"{field_name} '{value}' is not listed in {code_list_name}")

    return errors


async def run_opera_live_preflight(
    *,
    opera_url: str,
    evidence_dir: Path,
    allow_login: bool,
    authenticated_marker: str,
    headless: bool,
    browser_name: str,
) -> dict[str, Any]:
    if not opera_url:
        return {"status": "failed", "blocker": "opera_url is not configured"}

    username = os.getenv("OPERA_USERNAME")
    password = os.getenv("OPERA_PASSWORD")
    missing = missing_credential_names(username=username, password=password)
    evidence_dir.mkdir(parents=True, exist_ok=True)
    started = time.perf_counter()

    from playwright.async_api import async_playwright

    playwright = await async_playwright().start()
    browser = None
    try:
        browser_type = getattr(playwright, browser_name)
        browser = await browser_type.launch(headless=headless)
        context = await browser.new_context(
            locale="tr-TR",
            viewport={"width": 1440, "height": 1000},
        )
        page = await context.new_page()
        await page.goto(opera_url, wait_until="domcontentloaded", timeout=60000)
        try:
            await page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        await page.wait_for_timeout(1500)
        screenshot_path = evidence_dir / "opera_preflight_login.png"
        await page.screenshot(path=str(screenshot_path), full_page=True)
        interactive = await collect_interactive_elements(page)
        evidence = {
            "status": "blocked" if missing else "passed",
            "blocker": f"missing required secrets: {', '.join(missing)}" if missing else "",
            "current_url": page.url,
            "title": await page.title(),
            "screenshot": str(screenshot_path),
            "interactive_count": len(interactive),
            "interactive_elements": interactive[:40],
            "login_attempted": False,
            "duration_ms": round((time.perf_counter() - started) * 1000, 2),
        }
        if missing or not allow_login:
            if not missing and not allow_login:
                evidence["status"] = "blocked"
                evidence["blocker"] = "login is disabled; set opera_allow_login after selector review"
            return evidence

        login_result = await attempt_opera_login(
            page,
            username=username or "",
            password=password or "",
            authenticated_marker=authenticated_marker,
        )
        evidence.update(login_result)
        evidence["duration_ms"] = round((time.perf_counter() - started) * 1000, 2)
        return evidence
    except Exception as exc:
        return {
            "status": "failed",
            "blocker": sanitize_error(exc),
            "duration_ms": round((time.perf_counter() - started) * 1000, 2),
        }
    finally:
        if browser:
            await browser.close()
        await playwright.stop()


async def collect_interactive_elements(page: Any) -> list[dict[str, Any]]:
    return await page.locator("input, button, select, textarea, a").evaluate_all(
        """elements => elements.map((el, index) => ({
            index,
            tag: el.tagName.toLowerCase(),
            type: el.getAttribute('type') || '',
            name: el.getAttribute('name') || '',
            id: el.id || '',
            ariaLabel: el.getAttribute('aria-label') || '',
            text: (el.innerText || el.value || '').trim().slice(0, 80),
            placeholder: el.getAttribute('placeholder') || '',
            visible: !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)
        }))"""
    )


async def attempt_opera_login(page: Any, *, username: str, password: str, authenticated_marker: str = "") -> dict[str, Any]:
    username_locator = await first_available_locator(
        page,
        [
            "input[name='username']",
            "input[name='userid']",
            "input[type='email']",
            "input[type='text']",
            "input:not([type])",
        ],
    )
    password_locator = await first_available_locator(page, ["input[type='password']", "input[name='password']"])
    if username_locator is None or password_locator is None:
        return {
            "status": "failed",
            "blocker": "login fields were not discovered",
            "login_attempted": False,
        }
    await username_locator.fill(username)
    await password_locator.fill(password)
    submit = await first_available_locator(
        page,
        [
            "button[type='submit']",
            "input[type='submit']",
            "button:has-text('Sign In')",
            "button:has-text('Login')",
            "button:has-text('Log In')",
        ],
    )
    if submit is None:
        return {
            "status": "failed",
            "blocker": "login submit control was not discovered",
            "login_attempted": False,
        }
    await submit.click()
    try:
        await page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass
    await page.wait_for_timeout(2000)
    title = await page.title()
    password_visible = await locator_visible(password_locator)
    body_text = await safe_body_text(page)
    marker_visible = await authenticated_marker_visible(page, authenticated_marker)
    passed = login_reached_authenticated_state(
        current_url=page.url,
        password_visible=password_visible,
        body_text=body_text,
        authenticated_marker_visible=marker_visible,
    )
    return {
        "status": "passed" if passed else "blocked",
        "blocker": "" if passed else "login did not reach an authenticated OPERA page",
        "login_attempted": True,
        "post_login_url": page.url,
        "post_login_title": title,
    }


async def first_available_locator(page: Any, selectors: list[str]) -> Any | None:
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            if await locator.count() and await locator.is_visible(timeout=1000):
                return locator
        except Exception:
            continue
    return None


async def locator_visible(locator: Any) -> bool:
    try:
        return bool(await locator.is_visible(timeout=1000))
    except Exception:
        return False


async def safe_body_text(page: Any) -> str:
    try:
        return await page.locator("body").inner_text(timeout=1000)
    except Exception:
        return ""


async def authenticated_marker_visible(page: Any, marker: str) -> bool:
    if not marker:
        return False
    try:
        return bool(await page.locator(marker).first.is_visible(timeout=5000))
    except Exception:
        return False


def login_reached_authenticated_state(
    *,
    current_url: str,
    password_visible: bool,
    body_text: str,
    authenticated_marker_visible: bool,
) -> bool:
    text = body_text.lower()
    has_login_error = any(term in text for term in ("invalid", "incorrect", "denied", "hatal", "gecersiz"))
    return (
        "operacloud" in current_url.lower()
        and not password_visible
        and not has_login_error
        and authenticated_marker_visible
    )


def missing_credential_names(*, username: str | None, password: str | None) -> list[str]:
    missing = []
    if not username:
        missing.append("OPERA_USERNAME")
    if not password:
        missing.append("OPERA_PASSWORD")
    return missing


def build_execution_plan(fields: dict[str, Any]) -> list[dict[str, Any]]:
    plan = []
    for stage in STAGES:
        entry = dict(stage)
        if stage["stage"] == "CHILDREN":
            entry["enabled"] = int_or_zero(fields.get("children_count")) > 0
        else:
            entry["enabled"] = True
        plan.append(entry)
    return plan


def normalize_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def normalize_code_header(value: Any) -> str:
    normalized = normalize_label(value).replace(" ", "_")
    return normalized if normalized in {"market_codes", "source_codes", "rate_codes", "reservation_types", "room_types"} else normalized


def normalize_field_value(field_name: str, value: Any) -> Any:
    if field_name in {"arrival_date", "departure_date"}:
        return normalize_date(value)
    if field_name in {"rooms", "adults", "children_count"}:
        return int_or_zero(value)
    if field_name == "children_ages":
        return parse_children_ages(value)
    if value is None:
        return ""
    return str(value).strip()


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, ""):
        return ""
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def parse_iso_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return None


def int_or_zero(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return 0


def positive_decimal(value: Any) -> bool:
    if value in (None, ""):
        return False
    try:
        return float(str(value).strip()) > 0
    except ValueError:
        return False


def parse_children_ages(value: Any) -> list[int]:
    if value in (None, ""):
        return []
    if isinstance(value, (int, float)):
        return [int(value)]
    parts = re.split(r"[,;/\s]+", str(value).strip())
    return [int(part) for part in parts if part.isdigit()]


def reservation_record_id(fields: dict[str, Any]) -> str:
    guest = re.sub(r"[^A-Za-z0-9]+", "_", str(fields.get("guest_name") or "guest")).strip("_").lower()
    arrival = str(fields.get("arrival_date") or "arrival").replace("-", "")
    departure = str(fields.get("departure_date") or "departure").replace("-", "")
    return f"{guest}_{arrival}_{departure}"[:120]


def render_html_report(result: dict[str, Any]) -> str:
    rows = []
    for record in result["records"]:
        errors = "; ".join(record.get("validation_errors") or [])
        preflight = record.get("live_preflight") or {}
        screenshot = preflight.get("screenshot") or ""
        screenshot_link = f'<a href="{html.escape(screenshot)}">screenshot</a>' if screenshot else ""
        rows.append(
            "<tr>"
            f"<td>{html.escape(record['record_id'])}</td>"
            f"<td>{html.escape(record['status'])}</td>"
            f"<td>{html.escape(record.get('guest_name') or '')}</td>"
            f"<td>{html.escape(record.get('arrival_date') or '')}</td>"
            f"<td>{html.escape(record.get('departure_date') or '')}</td>"
            f"<td>{html.escape(record.get('room_type') or '')}</td>"
            f"<td>{html.escape(record.get('rate_code') or '')}</td>"
            f"<td>{html.escape(preflight.get('status') or '')} {screenshot_link}</td>"
            f"<td>{html.escape(errors)}</td>"
            "</tr>"
        )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Opera Rezervasyon Plan</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; color: #172033; background: #f7f8fb; }}
    header {{ padding: 24px 32px; background: #172033; color: white; }}
    h1 {{ margin: 0 0 6px; font-size: 24px; }}
    header p {{ margin: 0; color: #cbd5e1; }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 24px; }}
    table {{ width: 100%; border-collapse: collapse; background: white; border: 1px solid #d9dee8; }}
    th, td {{ padding: 9px 11px; border-bottom: 1px solid #d9dee8; text-align: left; vertical-align: top; }}
    th {{ background: #eef2f7; font-size: 12px; text-transform: uppercase; color: #46546a; }}
    td {{ overflow-wrap: anywhere; }}
  </style>
</head>
<body>
  <header>
    <h1>Opera Rezervasyon Plan</h1>
    <p>{html.escape(str(result['total_records']))} record(s), dry run only</p>
  </header>
  <main>
    <table>
      <thead><tr><th>Record</th><th>Status</th><th>Guest</th><th>Arrival</th><th>Departure</th><th>Room Type</th><th>Rate Code</th><th>Live Preflight</th><th>Validation</th></tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
  </main>
</body>
</html>"""


def bool_from_config(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def sanitize_error(error: object) -> str:
    return re.sub(r"\s+", " ", str(error or "")).strip()[:500]
