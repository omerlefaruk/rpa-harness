"""
Excel-driven RPA workflow: Trip.com recent reviews for OTA hotel rows.

Reads the OTA workbook, takes every Trip.com hotel link, derives the public
Trip.com review page, extracts reviews in the requested date window, and writes
JSON + XLSX outputs.
"""

from __future__ import annotations

import importlib.util
import json
import os
import re
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from harness import RPAWorkflow


ROOT = Path(__file__).resolve().parents[2]
REVIEW_MODULE_PATH = ROOT / "tests" / "browser" / "trip_marmara_taksim_reviews.py"
SPEC = importlib.util.spec_from_file_location("trip_marmara_taksim_reviews", REVIEW_MODULE_PATH)
review_tools = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
sys.modules[SPEC.name] = review_tools
SPEC.loader.exec_module(review_tools)


DEFAULT_INPUT = "data/ota_links.xlsx"
DEFAULT_OUTPUT_JSON = "runs/browser_recon/trip_com_all_hotels_recent_reviews.json"
DEFAULT_OUTPUT_XLSX = "reports/trip_com_all_hotels_recent_reviews.xlsx"
DEFAULT_SHEET = "Taksim Analiz"
TRIP_LINK_COLUMN = 7
HOTEL_NAME_COLUMN = 1
TRIP_SCORE_COLUMN = 5
TRIP_REVIEW_COUNT_COLUMN = 6


class TripComReviewsFromExcelWorkflow(RPAWorkflow):
    name = "trip_com_reviews_from_excel"
    tags = ["rpa", "excel", "trip", "reviews", "external"]
    max_retries_per_record = 1
    retry_base_delay_ms = 500

    async def setup(self):
        self.input_path = Path(self.config.variables.get("input_excel", DEFAULT_INPUT))
        self.sheet_name = self.config.variables.get("sheet", DEFAULT_SHEET)
        self.as_of = _as_of_date(self.config.variables.get("as_of_date"))
        self.start_date = self.as_of - timedelta(days=30)
        self.output_json = Path(self.config.variables.get("output_json", DEFAULT_OUTPUT_JSON))
        self.output_xlsx = Path(self.config.variables.get("output_excel", DEFAULT_OUTPUT_XLSX))
        self.records = read_trip_com_records(self.input_path, self.sheet_name)
        self.rows: list[dict] = []
        self.summary_rows: list[dict] = []
        self.log(
            f"Loaded {len(self.records)} Trip.com hotel records from "
            f"{self.input_path} / {self.sheet_name}"
        )

    def get_records(self):
        yield from self.records

    async def process_record(self, record: dict) -> dict:
        started = time.perf_counter()
        text, fetch_status = fetch_review_text_for_record(record)
        reviews = (
            review_tools.filter_recent_reviews(
                review_tools.extract_review_records(text, record["review_url"]),
                self.start_date,
                self.as_of,
            )
            if text
            else []
        )

        review_dicts = [review.to_dict() for review in reviews]
        self.rows.extend(
            {
                "hotel": record["hotel"],
                "hotel_id": record["hotel_id"],
                "source_row": record["source_row"],
                "review_url": record["review_url"],
                **review,
            }
            for review in review_dicts
        )
        self.summary_rows.append(
            {
                "hotel": record["hotel"],
                "hotel_id": record["hotel_id"],
                "source_row": record["source_row"],
                "trip_score": record.get("trip_score"),
                "trip_review_count": record.get("trip_review_count"),
                "review_url": record["review_url"],
                "status": fetch_status["status"],
                "attempts": fetch_status["attempts"],
                "text_length": fetch_status["text_length"],
                "recent_review_count": len(reviews),
                "duration_ms": round((time.perf_counter() - started) * 1000, 2),
                "error": fetch_status.get("error", ""),
            }
        )

        if fetch_status["status"] != "loaded":
            return {
                "status": "failed",
                "reason": f"Trip.com review page fetch failed: {fetch_status.get('error', '')}",
                "details": fetch_status,
            }
        return {
            "status": "passed",
            "details": {
                "hotel": record["hotel"],
                "hotel_id": record["hotel_id"],
                "recent_review_count": len(reviews),
                "attempts": fetch_status["attempts"],
            },
        }

    async def teardown(self):
        result = {
            "input_excel": str(self.input_path),
            "sheet": self.sheet_name,
            "last_30_days_window": {
                "start": self.start_date.isoformat(),
                "end": self.as_of.isoformat(),
            },
            "hotel_count": len(self.summary_rows),
            "total_recent_reviews": len(self.rows),
            "summary": self.summary_rows,
            "reviews": self.rows,
        }
        self.output_json.parent.mkdir(parents=True, exist_ok=True)
        self.output_json.write_text(
            json.dumps(result, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        write_results_workbook(
            self.output_xlsx,
            self.summary_rows,
            self.rows,
            result["last_30_days_window"],
        )
        self.result.output_files.extend([str(self.output_json), str(self.output_xlsx)])
        self.log(f"Wrote JSON: {self.output_json}")
        self.log(f"Wrote XLSX: {self.output_xlsx}")


def read_trip_com_records(input_path: Path, sheet_name: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(input_path, data_only=False)
    worksheet = workbook[sheet_name]
    records: list[dict] = []
    seen: set[str] = set()
    for row in range(3, worksheet.max_row + 1):
        hotel = worksheet.cell(row=row, column=HOTEL_NAME_COLUMN).value
        if not hotel:
            continue
        link_cell = worksheet.cell(row=row, column=TRIP_LINK_COLUMN)
        target = link_cell.hyperlink.target if link_cell.hyperlink else None
        if not target:
            continue
        hotel_id = extract_hotel_id(target)
        if not hotel_id or hotel_id in seen:
            continue
        seen.add(hotel_id)
        records.append(
            {
                "id": hotel_id,
                "hotel": str(hotel).strip(),
                "hotel_id": hotel_id,
                "source_row": row,
                "trip_score": worksheet.cell(row=row, column=TRIP_SCORE_COLUMN).value,
                "trip_review_count": worksheet.cell(row=row, column=TRIP_REVIEW_COUNT_COLUMN).value,
                "trip_url": target,
                "review_url": build_review_url(hotel_id, str(hotel)),
            }
        )
    return records


def extract_hotel_id(url: str) -> str | None:
    query_id = parse_qs(urlparse(url).query).get("hotelId")
    if query_id:
        return query_id[0]
    match = re.search(r"hotel-detail-(\d+)", url)
    return match.group(1) if match else None


def build_review_url(hotel_id: str, hotel_name: str) -> str:
    return f"https://in.trip.com/hotels/istanbul-hotel-detail-{hotel_id}/{slugify(hotel_name)}/review.html"


def slugify(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value.lower()).strip("-")


def fetch_review_text_for_record(record: dict) -> tuple[str, dict]:
    text, status = review_tools.fetch_public_text_with_status(record["review_url"])
    if text and review_tools.extract_review_records(text, record["review_url"]):
        return text, {
            "status": "loaded",
            "attempts": status["attempts"],
            "text_length": len(text),
        }
    if text:
        return text, {
            "status": "loaded_no_reviews",
            "attempts": status["attempts"],
            "text_length": len(text),
            "error": "Fetched page but no dated review records were parsed.",
        }
    return "", {
        "status": status["status"],
        "attempts": status["attempts"],
        "text_length": 0,
        "error": status.get("error", ""),
    }


def write_results_workbook(
    output_path: Path,
    summary_rows: list[dict],
    review_rows: list[dict],
    window: dict,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Window Start", window["start"], "Window End", window["end"]])
    summary.append([])
    summary_headers = [
        "Hotel",
        "Hotel ID",
        "Trip Score",
        "Trip Review Count",
        "Recent Review Count",
        "Status",
        "Attempts",
        "Review URL",
        "Error",
    ]
    summary.append(summary_headers)
    for row in summary_rows:
        summary.append(
            [
                row.get("hotel"),
                row.get("hotel_id"),
                row.get("trip_score"),
                row.get("trip_review_count"),
                row.get("recent_review_count"),
                row.get("status"),
                row.get("attempts"),
                row.get("review_url"),
                row.get("error"),
            ]
        )

    reviews = workbook.create_sheet("Reviews")
    review_headers = [
        "Hotel",
        "Hotel ID",
        "Date",
        "Reviewer",
        "Rating",
        "Label",
        "Room",
        "Traveler Type",
        "Review Text",
        "Property Response",
        "Source URL",
    ]
    reviews.append(review_headers)
    for row in review_rows:
        reviews.append(
            [
                row.get("hotel"),
                row.get("hotel_id"),
                row.get("date"),
                row.get("reviewer"),
                row.get("rating"),
                row.get("label"),
                row.get("room"),
                row.get("traveler_type"),
                row.get("text"),
                row.get("property_response"),
                row.get("source_url"),
            ]
        )

    for worksheet in workbook.worksheets:
        for cell in worksheet[3 if worksheet.title == "Summary" else 1]:
            cell.font = Font(bold=True)
        for column_cells in worksheet.columns:
            max_len = min(max(len(str(cell.value or "")) for cell in column_cells), 70)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, max_len + 2)
        worksheet.freeze_panes = "A4" if worksheet.title == "Summary" else "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)


def _as_of_date(configured: str | None = None) -> date:
    value = configured or os.getenv("TRIP_REVIEW_AS_OF_DATE")
    if value:
        return date.fromisoformat(value)
    return date.today()
