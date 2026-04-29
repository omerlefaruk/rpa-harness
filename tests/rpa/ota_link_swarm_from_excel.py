"""
Excel-driven RPA workflow: browser selector swarm for every OTA link.

Reads the OTA workbook, runs the browser selector swarm for every platform link
for every hotel, and writes an aggregate report. Individual link failures are
captured as output rows so the loop keeps going.
"""

from __future__ import annotations

import html
import json
import os
import re
import time
from pathlib import Path
from urllib.parse import urlparse

from harness import RPAWorkflow
from harness.selectors.browser_swarm import run_browser_selector_swarm


DEFAULT_INPUT = "data/ota_links.xlsx"
DEFAULT_SHEET = "Taksim Analiz"
DEFAULT_OUTPUT_DIR = "runs/ota_link_swarm"
DEFAULT_JSON = "runs/ota_link_swarm/ota_link_swarm_report.json"
DEFAULT_HTML = "reports/ota_link_swarm_report.html"
PLATFORM_START_COLUMN = 2
PLATFORM_WIDTH = 3
HOTEL_NAME_COLUMN = 1


class OtaLinkSwarmFromExcelWorkflow(RPAWorkflow):
    name = "ota_link_swarm_from_excel"
    tags = ["rpa", "excel", "browser", "swarm", "ota", "external"]
    max_retries_per_record = 0

    async def setup(self):
        self.input_path = Path(self.config.variables.get("input_excel", DEFAULT_INPUT))
        self.sheet_name = self.config.variables.get("sheet", DEFAULT_SHEET)
        self.output_dir = Path(self.config.variables.get("output_dir", DEFAULT_OUTPUT_DIR))
        self.output_json = Path(self.config.variables.get("output_json", DEFAULT_JSON))
        self.output_html = Path(self.config.variables.get("output_html", DEFAULT_HTML))
        self.max_candidates = int(self.config.variables.get("max_candidates", 15))
        self.timeout_ms = int(self.config.variables.get("timeout_ms", 20000))
        self.records = read_ota_link_records(self.input_path, self.sheet_name)
        self.rows: list[dict] = []
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.log(
            f"Loaded {len(self.records)} OTA links from "
            f"{self.input_path} / {self.sheet_name}"
        )

    def get_records(self):
        yield from self.records

    async def process_record(self, record: dict) -> dict:
        started = time.perf_counter()
        run_output = self.output_dir / safe_name(
            f"{record['source_row']:02d}_{record['hotel']}_{record['platform']}"
        )
        intent = (
            f"find hotel name, ratings, reviews, booking or review controls for "
            f"{record['hotel']} on {record['platform']}"
        )
        row = {
            **record,
            "status": "failed",
            "swarm_status": "not_run",
            "final_url": "",
            "title": "",
            "interactive_elements": 0,
            "candidates": 0,
            "validated": 0,
            "winner": None,
            "report": "",
            "html_report": "",
            "screenshot": "",
            "duration_ms": 0,
            "error": "",
        }
        try:
            report = await run_browser_selector_swarm(
                record["url"],
                output_dir=str(run_output),
                wait_until="domcontentloaded",
                timeout_ms=self.timeout_ms,
                max_candidates=self.max_candidates,
                intent=intent,
                use_subagents=False,
                save_raw_html=False,
            )
            row.update(
                {
                    "status": "processed",
                    "swarm_status": report.get("status", ""),
                    "final_url": report.get("url", ""),
                    "title": report.get("title", ""),
                    "interactive_elements": report.get("summary", {}).get("interactive_elements", 0),
                    "candidates": report.get("summary", {}).get("candidates", 0),
                    "validated": report.get("summary", {}).get("validated", 0),
                    "winner": report.get("validation", {}).get("winner"),
                    "report": report.get("artifacts", {}).get("report", ""),
                    "html_report": report.get("artifacts", {}).get("html_report", ""),
                    "screenshot": report.get("artifacts", {}).get("screenshot", ""),
                }
            )
        except Exception as exc:
            row.update(
                {
                    "status": "error",
                    "swarm_status": "error",
                    "error": sanitize_error(exc),
                }
            )
        finally:
            row["duration_ms"] = round((time.perf_counter() - started) * 1000, 2)
            self.rows.append(row)

        return {
            "status": "passed",
            "details": {
                "hotel": record["hotel"],
                "platform": record["platform"],
                "swarm_status": row["swarm_status"],
                "error": row["error"],
            },
        }

    async def teardown(self):
        result = {
            "input_excel": str(self.input_path),
            "sheet": self.sheet_name,
            "total_links": len(self.rows),
            "processed": sum(1 for row in self.rows if row["status"] == "processed"),
            "errors": sum(1 for row in self.rows if row["status"] == "error"),
            "passed_swarm": sum(1 for row in self.rows if row["swarm_status"] == "passed"),
            "no_winner": sum(1 for row in self.rows if row["swarm_status"] == "no_winner"),
            "rows": self.rows,
        }
        self.output_json.parent.mkdir(parents=True, exist_ok=True)
        self.output_html.parent.mkdir(parents=True, exist_ok=True)
        self.output_json.write_text(
            json.dumps(result, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        self.output_html.write_text(
            render_html_report(result, result_path=str(self.output_html)),
            encoding="utf-8",
        )
        self.result.output_files.extend([str(self.output_json), str(self.output_html)])
        self.log(f"Wrote JSON: {self.output_json}")
        self.log(f"Wrote HTML: {self.output_html}")


def read_ota_link_records(input_path: Path, sheet_name: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(input_path, data_only=False)
    worksheet = workbook[sheet_name]
    platforms = platform_link_columns(worksheet)
    records: list[dict] = []
    for row in range(3, worksheet.max_row + 1):
        hotel = worksheet.cell(row=row, column=HOTEL_NAME_COLUMN).value
        if not hotel:
            continue
        for platform, link_column in platforms:
            link_cell = worksheet.cell(row=row, column=link_column)
            target = link_cell.hyperlink.target if link_cell.hyperlink else None
            if not target:
                continue
            records.append(
                {
                    "id": f"{row}:{platform}",
                    "source_row": row,
                    "hotel": str(hotel).strip(),
                    "platform": str(platform).strip(),
                    "url": target,
                    "domain": urlparse(target).netloc,
                }
            )
    return records


def platform_link_columns(worksheet) -> list[tuple[str, int]]:
    columns: list[tuple[str, int]] = []
    col = PLATFORM_START_COLUMN
    while col <= worksheet.max_column:
        platform = worksheet.cell(row=1, column=col).value
        link_col = col + PLATFORM_WIDTH - 1
        if platform and str(platform).strip().lower() != "ortalama":
            columns.append((str(platform), link_col))
        col += PLATFORM_WIDTH
    return columns


def render_html_report(result: dict, *, result_path: str = DEFAULT_HTML) -> str:
    rows = result["rows"]
    table_rows = []
    sections = []
    for row in rows:
        html_report = row.get("html_report") or ""
        screenshot = row.get("screenshot") or ""
        report_link = (
            f'<a href="{html.escape(relative_or_raw(html_report, result_path=result_path))}">Swarm report</a>'
            if html_report
            else ""
        )
        screenshot_link = (
            f'<a href="{html.escape(relative_or_raw(screenshot, result_path=result_path))}">Screenshot</a>'
            if screenshot
            else ""
        )
        table_rows.append(
            "<tr>"
            f"<td>{html.escape(row['hotel'])}</td>"
            f"<td>{html.escape(row['platform'])}</td>"
            f"<td><span class=\"badge {html.escape(row['swarm_status'])}\">{html.escape(row['swarm_status'])}</span></td>"
            f"<td class=\"num\">{html.escape(str(row['interactive_elements']))}</td>"
            f"<td class=\"num\">{html.escape(str(row['validated']))}</td>"
            f"<td>{html.escape(row.get('domain', ''))}</td>"
            f"<td>{report_link} {screenshot_link}</td>"
            "</tr>"
        )
        winner = row.get("winner") or {}
        selector = winner.get("selector") if isinstance(winner, dict) else None
        sections.append(
            f"""
            <article class="card">
              <h2>{html.escape(row['hotel'])} <span>{html.escape(row['platform'])}</span></h2>
              <dl>
                <dt>Status</dt><dd>{html.escape(row['swarm_status'])}</dd>
                <dt>URL</dt><dd><a href="{html.escape(row['url'])}">{html.escape(row['url'])}</a></dd>
                <dt>Final URL</dt><dd>{html.escape(row.get('final_url') or '')}</dd>
                <dt>Title</dt><dd>{html.escape(row.get('title') or '')}</dd>
                <dt>Winner</dt><dd>{html.escape(json.dumps(selector, ensure_ascii=False) if selector else '')}</dd>
                <dt>Error</dt><dd>{html.escape(row.get('error') or '')}</dd>
                <dt>Artifacts</dt><dd>{report_link} {screenshot_link}</dd>
              </dl>
            </article>
            """
        )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>OTA Link Swarm Report</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; color: #182033; background: #f5f7fb; }}
    header {{ background: #101828; color: white; padding: 28px 36px; }}
    h1 {{ margin: 0 0 8px; font-size: 28px; }}
    header p {{ margin: 0; color: #cbd5e1; }}
    main {{ max-width: 1220px; margin: 0 auto; padding: 24px; }}
    .metrics {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 18px; }}
    .metric, .table-wrap, .card {{ background: white; border: 1px solid #d9dee8; border-radius: 8px; }}
    .metric {{ padding: 16px; }}
    .metric b {{ display: block; font-size: 24px; }}
    .metric span {{ color: #647084; font-size: 13px; }}
    .table-wrap {{ overflow-x: auto; margin-bottom: 18px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 9px 11px; border-bottom: 1px solid #d9dee8; text-align: left; vertical-align: top; }}
    th {{ background: #eef2f7; font-size: 12px; color: #46546a; text-transform: uppercase; }}
    .num {{ text-align: right; }}
    .badge {{ display: inline-block; border-radius: 999px; padding: 3px 8px; font-size: 12px; background: #eef2f7; color: #334155; }}
    .badge.passed {{ background: #e7f7ee; color: #147a46; }}
    .badge.error {{ background: #feeceb; color: #b42318; }}
    .card {{ padding: 16px; margin: 12px 0; }}
    .card h2 {{ margin: 0 0 12px; font-size: 18px; }}
    .card h2 span {{ color: #2457d6; font-size: 14px; margin-left: 8px; }}
    dl {{ display: grid; grid-template-columns: 130px minmax(0, 1fr); gap: 8px 12px; margin: 0; }}
    dt {{ color: #647084; font-weight: 700; }}
    dd {{ margin: 0; overflow-wrap: anywhere; }}
    a {{ color: #2457d6; }}
    @media (max-width: 760px) {{ .metrics {{ grid-template-columns: 1fr; }} dl {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <header>
    <h1>OTA Link Swarm Report</h1>
    <p>{html.escape(str(result['total_links']))} links from {html.escape(result['input_excel'])} / {html.escape(result['sheet'])}</p>
  </header>
  <main>
    <section class="metrics">
      <div class="metric"><b>{result['total_links']}</b><span>Total links</span></div>
      <div class="metric"><b>{result['processed']}</b><span>Processed</span></div>
      <div class="metric"><b>{result['passed_swarm']}</b><span>Swarm winners</span></div>
      <div class="metric"><b>{result['errors']}</b><span>Errors captured</span></div>
    </section>
    <section class="table-wrap">
      <table>
        <thead><tr><th>Hotel</th><th>Platform</th><th>Status</th><th class="num">Elements</th><th class="num">Validated</th><th>Domain</th><th>Artifacts</th></tr></thead>
        <tbody>{''.join(table_rows)}</tbody>
      </table>
    </section>
    {''.join(sections)}
  </main>
</body>
</html>"""


def relative_or_raw(path: str, *, result_path: str) -> str:
    try:
        return os.path.relpath(Path(path).resolve(), Path(result_path).resolve().parent)
    except Exception:
        return path


def safe_name(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_.-]+", "_", value).strip("_")[:120]


def sanitize_error(error: object) -> str:
    return re.sub(r"\s+", " ", str(error or "")).strip()[:500]
