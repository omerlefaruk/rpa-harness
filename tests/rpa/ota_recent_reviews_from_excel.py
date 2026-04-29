"""
Excel-driven RPA workflow: recent OTA reviews for every hotel link.

Reads the OTA workbook, loops every website link for every hotel, extracts
recent reviews when the site exposes parseable review text, and writes a report
for all links including blocked and empty outcomes.
"""

from __future__ import annotations

import html
import importlib.util
import json
import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

from harness import RPAWorkflow


ROOT = Path(__file__).resolve().parents[2]


def _load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


ota_links = _load_module("ota_link_swarm_from_excel", ROOT / "tests" / "rpa" / "ota_link_swarm_from_excel.py")
trip_excel = _load_module("trip_com_reviews_from_excel", ROOT / "tests" / "rpa" / "trip_com_reviews_from_excel.py")
trip_reviews = trip_excel.review_tools


DEFAULT_INPUT = "data/ota_links.xlsx"
DEFAULT_SHEET = "Taksim Analiz"
DEFAULT_JSON = "runs/ota_recent_reviews/ota_recent_reviews_report.json"
DEFAULT_HTML = "reports/ota_recent_reviews_report.html"
DEFAULT_XLSX = "reports/ota_recent_reviews_report.xlsx"
PAGE_TIMEOUT_MS = 20000
BODY_TEXT_TIMEOUT_MS = 5000
MAX_TEXT_ATTEMPTS = 2

EN_MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}
TR_MONTHS = trip_reviews.TR_MONTHS

DATE_PATTERNS = [
    re.compile(
        r"\b(?P<month>Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
        r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t|tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
        r"\s+(?P<day>\d{1,2}),?\s+(?P<year>20\d{2})\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b(?P<day>\d{1,2})\s+"
        r"(?P<month>Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
        r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t|tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
        r"\s+(?P<year>20\d{2})\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b(?P<day>\d{1,2})\s+"
        r"(?P<month>Ocak|Şubat|Mart|Nisan|Mayıs|Haziran|Temmuz|Ağustos|Eylül|Ekim|Kasım|Aralık)"
        r"\s+(?P<year>20\d{2})\b",
    ),
    re.compile(r"\b(?P<year>20\d{2})-(?P<month>\d{1,2})-(?P<day>\d{1,2})\b"),
]

RELATIVE_PATTERNS = [
    re.compile(r"\b(?P<count>\d+)\s+(?P<unit>day|days|week|weeks)\s+ago\b", re.IGNORECASE),
    re.compile(r"\b(?P<count_word>a|an)\s+(?P<unit>day|week|month)\s+ago\b", re.IGNORECASE),
    re.compile(r"\b(?P<count>\d+)\s+(?P<unit>month|months)\s+ago\b", re.IGNORECASE),
    re.compile(r"\b(?P<count>\d+)\s+(?P<unit>gün|hafta)\s+önce\b", re.IGNORECASE),
    re.compile(r"\b(?P<word>yesterday|dün)\b", re.IGNORECASE),
]

REVIEW_KEYWORDS = [
    "review",
    "reviews",
    "guest",
    "traveler",
    "traveller",
    "rating",
    "rated",
    "yorum",
    "yorumlar",
    "değerlendirme",
    "degerlendirme",
    "puan",
    "misafir",
    "konuk",
]

BLOCKED_MARKERS = [
    "access denied",
    "too many requests",
    "bot or not",
    "show us your human side",
    "can't tell if you're a human or a bot",
    "checking your browser",
    "verify you are human",
    "unusual traffic",
    "google.com/sorry",
    "captcha",
]
GOOGLE_REVIEW_RE = re.compile(
    r"\s+(?P<rating>\d(?:\.\d)?/5)\s+"
    r"(?P<relative>(?:a|an|\d+)\s+(?:day|days|week|weeks|month|months)\s+ago|yesterday)"
    r"\s+on\s+(?P<source>Google|Trip\.com|Agoda|Booking\.com|Expedia|Hotels\.com)\s+",
    re.IGNORECASE,
)


class ParsedReview:
    def __init__(
        self,
        *,
        date: date,
        reviewer: str,
        rating: str,
        title: str,
        text: str,
        source_url: str,
        extraction_method: str,
    ):
        self.date = date
        self.reviewer = reviewer
        self.rating = rating
        self.title = title
        self.text = text
        self.source_url = source_url
        self.extraction_method = extraction_method

    date: date
    reviewer: str
    rating: str
    title: str
    text: str
    source_url: str
    extraction_method: str

    def to_dict(self) -> dict:
        return {
            "date": self.date.isoformat(),
            "reviewer": self.reviewer,
            "rating": self.rating,
            "title": self.title,
            "text": self.text,
            "source_url": self.source_url,
            "extraction_method": self.extraction_method,
        }


class OtaRecentReviewsFromExcelWorkflow(RPAWorkflow):
    name = "ota_recent_reviews_from_excel"
    tags = ["rpa", "excel", "browser", "reviews", "ota", "external"]
    max_retries_per_record = 0

    async def setup(self):
        self.input_path = Path(self.config.variables.get("input_excel", DEFAULT_INPUT))
        self.sheet_name = self.config.variables.get("sheet", DEFAULT_SHEET)
        self.as_of = _as_of_date(self.config.variables.get("as_of_date"))
        self.start_date = self.as_of - timedelta(days=30)
        self.output_json = Path(self.config.variables.get("output_json", DEFAULT_JSON))
        self.output_html = Path(self.config.variables.get("output_html", DEFAULT_HTML))
        self.output_xlsx = Path(self.config.variables.get("output_excel", DEFAULT_XLSX))
        self.timeout_ms = int(self.config.variables.get("timeout_ms", PAGE_TIMEOUT_MS))
        self.records = ota_links.read_ota_link_records(self.input_path, self.sheet_name)
        self.summary_rows: list[dict] = []
        self.review_rows: list[dict] = []
        self.playwright = None
        self.browser = None
        self.page = None
        self.log(
            f"Loaded {len(self.records)} OTA links from "
            f"{self.input_path} / {self.sheet_name}; review window "
            f"{self.start_date.isoformat()} to {self.as_of.isoformat()}"
        )

    def get_records(self):
        yield from self.records

    async def process_record(self, record: dict) -> dict:
        started = time.perf_counter()
        summary = {
            **record,
            "status": "not_run",
            "fetch_status": "",
            "fetch_method": "",
            "final_url": record["url"],
            "title": "",
            "text_length": 0,
            "recent_review_count": 0,
            "blocked": False,
            "duration_ms": 0,
            "error": "",
        }
        reviews: list[ParsedReview] = []
        try:
            if is_trip_com(record):
                reviews, fetch_status = collect_trip_com_reviews(record, self.start_date, self.as_of)
            elif is_google(record):
                text, fetch_status = await self.fetch_google_maps_reviews(record)
                reviews = extract_google_recent_reviews(
                    text,
                    fetch_status.get("final_url") or record["url"],
                    start_date=self.start_date,
                    end_date=self.as_of,
                )
            elif is_yandex(record):
                text, fetch_status = await self.fetch_yandex_maps_reviews(record)
                reviews = extract_generic_recent_reviews(
                    text,
                    fetch_status.get("final_url") or record["url"],
                    start_date=self.start_date,
                    end_date=self.as_of,
                )
            else:
                text, fetch_status = await self.fetch_text(record)
                reviews = extract_generic_recent_reviews(
                    text,
                    record["url"],
                    start_date=self.start_date,
                    end_date=self.as_of,
                )
            summary.update(fetch_status)
            summary["fetch_status"] = fetch_status.get("status", "")
            summary["recent_review_count"] = len(reviews)
            summary["status"] = "processed"
        except Exception as exc:
            summary.update(
                {
                    "status": "error",
                    "blocked": False,
                    "error": sanitize_error(exc),
                }
            )
        finally:
            summary["duration_ms"] = round((time.perf_counter() - started) * 1000, 2)
            self.summary_rows.append(summary)
            self.review_rows.extend(
                {
                    "hotel": record["hotel"],
                    "platform": record["platform"],
                    "domain": record["domain"],
                    "source_row": record["source_row"],
                    **review.to_dict(),
                }
                for review in reviews
            )

        return {
            "status": "passed",
            "details": {
                "hotel": record["hotel"],
                "platform": record["platform"],
                "recent_review_count": summary["recent_review_count"],
                "status": summary["status"],
                "error": summary["error"],
            },
        }

    async def fetch_text(self, record: dict) -> tuple[str, dict]:
        text, http_status = trip_reviews.fetch_public_text_with_status(record["url"])
        if text and not looks_blocked(text):
            return text, {
                "status": "loaded",
                "fetch_method": "http_public_html",
                "final_url": record["url"],
                "title": "",
                "text_length": len(text),
                "blocked": False,
                "error": "",
            }

        browser_text, browser_status = await self.fetch_text_with_browser(record["url"])
        if browser_text:
            return browser_text, browser_status

        return "", {
            "status": "blocked_or_empty" if looks_blocked(text) else http_status["status"],
            "fetch_method": "http_public_html",
            "final_url": record["url"],
            "title": "",
            "text_length": len(text),
            "blocked": looks_blocked(text),
            "error": http_status.get("error") or ("blocked or empty page" if not text else ""),
        }

    async def fetch_text_with_browser(self, url: str) -> tuple[str, dict]:
        await self.ensure_browser()
        last_error = None
        for attempt in range(1, MAX_TEXT_ATTEMPTS + 1):
            try:
                await self.page.goto(url, wait_until="domcontentloaded", timeout=self.timeout_ms)
                await self.page.wait_for_timeout(2000)
                title = await self.page.title()
                final_url = self.page.url
                text = await self.page.locator("body").inner_text(timeout=BODY_TEXT_TIMEOUT_MS)
                text = normalize_text(text)
                return text, {
                    "status": "blocked_or_empty" if looks_blocked(text + " " + final_url) else "loaded",
                    "fetch_method": "playwright_body_text",
                    "final_url": final_url,
                    "title": title,
                    "text_length": len(text),
                    "blocked": looks_blocked(text + " " + final_url),
                    "error": "",
                }
            except Exception as exc:
                last_error = exc
                await self.page.wait_for_timeout(500 * attempt)
        return "", {
            "status": "failed",
            "fetch_method": "playwright_body_text",
            "final_url": url,
            "title": "",
            "text_length": 0,
            "blocked": False,
            "error": sanitize_error(last_error),
        }

    async def fetch_google_maps_reviews(self, record: dict) -> tuple[str, dict]:
        await self.ensure_browser()
        search_url = f"https://www.google.com/maps/search/?api=1&query={quote(record['hotel'])}&hl=en"
        last_error = None
        for attempt in range(1, MAX_TEXT_ATTEMPTS + 1):
            try:
                await self.page.goto(search_url, wait_until="domcontentloaded", timeout=self.timeout_ms)
                await self.page.wait_for_timeout(6000)
                reviews_url = google_maps_reviews_url(self.page.url)
                await self.page.goto(reviews_url, wait_until="domcontentloaded", timeout=self.timeout_ms)
                await self.page.wait_for_timeout(6000)
                await self.click_google_newest_sort()
                await self.scroll_google_reviews()
                title = await self.page.title()
                final_url = self.page.url
                text = normalize_text(await self.page.locator("body").inner_text(timeout=BODY_TEXT_TIMEOUT_MS))
                return text, {
                    "status": "loaded",
                    "fetch_method": "google_maps_reviews",
                    "final_url": final_url,
                    "title": title,
                    "text_length": len(text),
                    "blocked": looks_blocked(text + " " + final_url),
                    "error": "",
                }
            except Exception as exc:
                last_error = exc
                await self.page.wait_for_timeout(700 * attempt)
        return "", {
            "status": "failed",
            "fetch_method": "google_maps_reviews",
            "final_url": search_url,
            "title": "",
            "text_length": 0,
            "blocked": False,
            "error": sanitize_error(last_error),
        }

    async def click_google_newest_sort(self) -> None:
        try:
            sort_button = self.page.get_by_text("Most relevant", exact=False).first
            if await sort_button.count() and await sort_button.is_visible(timeout=1000):
                await sort_button.click(timeout=2000)
                await self.page.wait_for_timeout(1000)
                newest = self.page.get_by_text("Newest", exact=True).first
                if await newest.count() and await newest.is_visible(timeout=1000):
                    await newest.click(timeout=2000)
                    await self.page.wait_for_timeout(3000)
        except Exception:
            return

    async def scroll_google_reviews(self) -> None:
        for _ in range(8):
            try:
                await self.page.mouse.wheel(0, 1800)
            except Exception:
                await self.page.evaluate("() => window.scrollBy(0, 1800)")
            await self.page.wait_for_timeout(700)

    async def fetch_yandex_maps_reviews(self, record: dict) -> tuple[str, dict]:
        await self.ensure_browser()
        search_url = f"https://yandex.com.tr/maps/?text={quote(record['hotel'])}&source=serp_navig"
        last_error = None
        for attempt in range(1, MAX_TEXT_ATTEMPTS + 1):
            try:
                await self.page.goto(search_url, wait_until="domcontentloaded", timeout=self.timeout_ms)
                await self.page.wait_for_timeout(6000)
                reviews_url = yandex_maps_reviews_url(self.page.url)
                await self.page.goto(reviews_url, wait_until="domcontentloaded", timeout=self.timeout_ms)
                await self.page.wait_for_timeout(6000)
                await self.scroll_google_reviews()
                title = await self.page.title()
                final_url = self.page.url
                text = normalize_text(await self.page.locator("body").inner_text(timeout=BODY_TEXT_TIMEOUT_MS))
                return text, {
                    "status": "loaded",
                    "fetch_method": "yandex_maps_reviews",
                    "final_url": final_url,
                    "title": title,
                    "text_length": len(text),
                    "blocked": looks_blocked(text + " " + final_url),
                    "error": "",
                }
            except Exception as exc:
                last_error = exc
                await self.page.wait_for_timeout(700 * attempt)
        return "", {
            "status": "failed",
            "fetch_method": "yandex_maps_reviews",
            "final_url": search_url,
            "title": "",
            "text_length": 0,
            "blocked": False,
            "error": sanitize_error(last_error),
        }

    async def ensure_browser(self) -> None:
        if self.page:
            return
        from playwright.async_api import async_playwright

        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=True)
        context = await self.browser.new_context(
            locale="tr-TR",
            user_agent=trip_reviews.public_request_headers()["User-Agent"],
        )
        self.page = await context.new_page()

    async def teardown(self):
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()

        result = {
            "input_excel": str(self.input_path),
            "sheet": self.sheet_name,
            "last_30_days_window": {
                "start": self.start_date.isoformat(),
                "end": self.as_of.isoformat(),
            },
            "total_links": len(self.summary_rows),
            "processed": sum(1 for row in self.summary_rows if row["status"] == "processed"),
            "errors": sum(1 for row in self.summary_rows if row["status"] == "error"),
            "blocked_or_empty": sum(1 for row in self.summary_rows if row.get("blocked")),
            "failed_or_blocked": sum(
                1
                for row in self.summary_rows
                if row.get("blocked")
                or row.get("fetch_status") in {"failed", "blocked_or_empty"}
                or row.get("status") == "error"
            ),
            "links_with_recent_reviews": sum(1 for row in self.summary_rows if row["recent_review_count"] > 0),
            "total_recent_reviews": len(self.review_rows),
            "summary": self.summary_rows,
            "reviews": self.review_rows,
        }
        self.output_json.parent.mkdir(parents=True, exist_ok=True)
        self.output_html.parent.mkdir(parents=True, exist_ok=True)
        self.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        self.output_json.write_text(
            json.dumps(result, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        self.output_html.write_text(
            render_html_report(result, result_path=str(self.output_html)),
            encoding="utf-8",
        )
        write_results_workbook(self.output_xlsx, result)
        self.result.output_files.extend(
            [str(self.output_json), str(self.output_html), str(self.output_xlsx)]
        )
        self.log(f"Wrote JSON: {self.output_json}")
        self.log(f"Wrote HTML: {self.output_html}")
        self.log(f"Wrote XLSX: {self.output_xlsx}")


def is_trip_com(record: dict) -> bool:
    return "trip.com" in record.get("domain", "") or str(record.get("platform", "")).lower() == "trip.com"


def is_google(record: dict) -> bool:
    return "google." in record.get("domain", "") or str(record.get("platform", "")).lower() == "google"


def is_yandex(record: dict) -> bool:
    return "yandex." in record.get("domain", "") or str(record.get("platform", "")).lower() == "yandex"


def google_maps_reviews_url(url: str) -> str:
    reviews_url = url.replace("!4m9!3m8", "!4m11!3m10")
    if "!9m1!1b1" not in reviews_url and "!16s" in reviews_url:
        reviews_url = reviews_url.replace("!16s", "!9m1!1b1!16s", 1)
    return reviews_url


def yandex_maps_reviews_url(url: str) -> str:
    if "/maps/org/" not in url:
        return url
    base, _, query = url.partition("?")
    base = base.rstrip("/")
    if not base.endswith("/reviews"):
        base = f"{base}/reviews"
    separator = "?" if query else ""
    suffix = f"{separator}{query}" if query else ""
    if "tab=reviews" not in suffix:
        suffix = f"{suffix}{'&' if suffix else '?'}tab=reviews"
    return f"{base}/{suffix}"


def collect_trip_com_reviews(record: dict, start_date: date, end_date: date) -> tuple[list[ParsedReview], dict]:
    hotel_id = trip_excel.extract_hotel_id(record["url"])
    if not hotel_id:
        return [], {
            "status": "failed",
            "fetch_method": "trip_com_public_review_page",
            "final_url": record["url"],
            "title": "",
            "text_length": 0,
            "blocked": False,
            "error": "Trip.com hotelId was not found in URL.",
        }
    trip_record = {
        **record,
        "hotel_id": hotel_id,
        "review_url": trip_excel.build_review_url(hotel_id, record["hotel"]),
    }
    text, fetch_status = trip_excel.fetch_review_text_for_record(trip_record)
    raw_reviews = (
        trip_reviews.filter_recent_reviews(
            trip_reviews.extract_review_records(text, trip_record["review_url"]),
            start_date,
            end_date,
        )
        if text
        else []
    )
    reviews = [
        ParsedReview(
            date=review.date,
            reviewer=review.reviewer or "",
            rating=review.rating or "",
            title=review.label or "",
            text=review.text,
            source_url=review.source_url,
            extraction_method="trip_com_public_review_page",
        )
        for review in raw_reviews
    ]
    return reviews, {
        "status": fetch_status["status"],
        "fetch_method": "trip_com_public_review_page",
        "final_url": trip_record["review_url"],
        "title": "",
        "text_length": fetch_status.get("text_length", 0),
        "blocked": False,
        "error": fetch_status.get("error", ""),
    }


def extract_generic_recent_reviews(
    text: str,
    source_url: str,
    *,
    start_date: date,
    end_date: date,
) -> list[ParsedReview]:
    if not text:
        return []
    normalized = normalize_text(text)
    candidates: list[ParsedReview] = []
    for review_date, start, end, method in find_dates(normalized, end_date):
        if not start_date <= review_date <= end_date:
            continue
        left = max(0, start - 420)
        right = min(len(normalized), end + 900)
        snippet = clean_review_snippet(normalized[left:right], date_offset=start - left)
        if not looks_like_review_snippet(snippet):
            continue
        candidates.append(
            ParsedReview(
                date=review_date,
                reviewer=infer_generic_reviewer(snippet),
                rating=infer_generic_rating(snippet),
                title="",
                text=snippet[:1800],
                source_url=source_url,
                extraction_method=method,
            )
        )
    return dedupe_parsed_reviews(candidates)


def extract_google_recent_reviews(
    text: str,
    source_url: str,
    *,
    start_date: date,
    end_date: date,
) -> list[ParsedReview]:
    normalized = normalize_text(text)
    matches = list(GOOGLE_REVIEW_RE.finditer(normalized))
    records: list[ParsedReview] = []
    for index, match in enumerate(matches):
        if match.group("source").lower() != "google":
            continue
        review_date = parse_relative_date_text(match.group("relative"), end_date)
        if not review_date or not start_date <= review_date <= end_date:
            continue
        next_start = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        body = trim_google_review_body(normalized[match.end():next_start])
        if len(body) < 20:
            continue
        records.append(
            ParsedReview(
                date=review_date,
                reviewer=clean_google_reviewer(normalized[max(0, match.start() - 180):match.start()]),
                rating=match.group("rating"),
                title="",
                text=body,
                source_url=source_url,
                extraction_method="google_maps_reviews",
            )
        )
    return dedupe_parsed_reviews(records)


def clean_google_reviewer(value: str) -> str:
    cleaned = normalize_text(value)
    if "" in cleaned:
        cleaned = cleaned.rsplit("", 1)[-1]
        cleaned = re.sub(
            r"^\s*\d(?:\.\d)?/5\s+(?:a|an|\d+)\s+(?:day|days|week|weeks|month|months)\s+ago\s+on\s+\S+\s+",
            "",
            cleaned,
            flags=re.IGNORECASE,
        )
    for marker in [" More ", " Like ", " Share ", "All reviews", "Most relevant"]:
        if marker in cleaned:
            cleaned = cleaned.rsplit(marker, 1)[-1]
    cleaned = re.sub(r"\bLocal Guide\b.*$", "", cleaned).strip()
    cleaned = re.sub(r"\b\d+\s+reviews?.*$", "", cleaned).strip()
    parts = [part.strip() for part in re.split(r"[.!?]\s+", cleaned) if part.strip()]
    if parts:
        cleaned = parts[-1]
    cleaned = re.sub(r"^[^\w]+", "", cleaned).strip()
    return cleaned[-80:].strip()


def trim_google_review_body(value: str) -> str:
    cleaned = normalize_text(value)
    cleaned = re.split(r"\bResponse from the owner\b", cleaned)[0]
    cleaned = re.split(r"\s+\s+", cleaned)[0]
    cleaned = re.split(r"\s+Like\s+", cleaned)[0]
    cleaned = re.split(r"\s+Share\s+", cleaned)[0]
    cleaned = cleaned.replace("… More", "").replace("Read more", "").strip()
    return cleaned[:1800]


def find_dates(text: str, as_of: date) -> list[tuple[date, int, int, str]]:
    dates: list[tuple[date, int, int, str]] = []
    for pattern in DATE_PATTERNS:
        for match in pattern.finditer(text):
            if is_property_response_date(text, match.start()):
                continue
            parsed = parse_date_match(match)
            if parsed:
                dates.append((parsed, match.start(), match.end(), "generic_absolute_date"))
    for pattern in RELATIVE_PATTERNS:
        for match in pattern.finditer(text):
            parsed = parse_relative_date_match(match, as_of)
            if parsed:
                dates.append((parsed, match.start(), match.end(), "generic_relative_date"))
    return sorted(dates, key=lambda item: item[1])


def is_property_response_date(text: str, match_start: int) -> bool:
    prefix = text[max(0, match_start - 50):match_start].lower()
    markers = [
        "cevap verilme tarihi",
        "yanıt tarihi",
        "yanit tarihi",
        "response date",
        "responded",
        "response from property",
    ]
    return any(marker in prefix for marker in markers)


def parse_date_match(match: re.Match) -> date | None:
    groups = match.groupdict()
    try:
        year = int(groups["year"])
        day = int(groups["day"])
        month_value = groups["month"]
        if month_value.isdigit():
            month = int(month_value)
        elif month_value in TR_MONTHS:
            month = TR_MONTHS[month_value]
        else:
            month = EN_MONTHS[month_value.lower().rstrip(".")]
        return date(year, month, day)
    except (KeyError, TypeError, ValueError):
        return None


def parse_relative_date_match(match: re.Match, as_of: date) -> date | None:
    groups = match.groupdict()
    if groups.get("word"):
        return as_of - timedelta(days=1)
    try:
        count = 1 if groups.get("count_word") else int(groups["count"])
    except (KeyError, TypeError, ValueError):
        return None
    unit = groups.get("unit", "").lower()
    if unit in {"month", "months"}:
        return as_of - timedelta(days=count * 30)
    if unit in {"week", "weeks", "hafta"}:
        return as_of - timedelta(days=count * 7)
    if unit in {"day", "days", "gün"}:
        return as_of - timedelta(days=count)
    return None


def parse_relative_date_text(value: str, as_of: date) -> date | None:
    cleaned = normalize_text(value).lower()
    if cleaned == "yesterday":
        return as_of - timedelta(days=1)
    match = re.match(r"(?P<count>\d+|a|an)\s+(?P<unit>day|days|week|weeks|month|months)\s+ago", cleaned)
    if not match:
        return None
    count_text = match.group("count")
    count = 1 if count_text in {"a", "an"} else int(count_text)
    unit = match.group("unit")
    if unit.startswith("day"):
        return as_of - timedelta(days=count)
    if unit.startswith("week"):
        return as_of - timedelta(days=count * 7)
    if unit.startswith("month"):
        return as_of - timedelta(days=count * 30)
    return None


def clean_review_snippet(value: str, *, date_offset: int | None = None) -> str:
    cleaned = normalize_text(value)
    if date_offset is not None:
        date_offset = max(0, min(date_offset, len(cleaned)))
        prefix = cleaned[:date_offset]
        boundary_markers = [
            "Bu değerlendirmeyi yararlı buldunuz mu?",
            "Bu değerlendirmeyi yararlı buldu",
            "EVET|HAYIR",
            "Translate",
            "Dilime çevir",
        ]
        best = -1
        best_marker = ""
        for marker in boundary_markers:
            idx = prefix.rfind(marker)
            if idx > best:
                best = idx
                best_marker = marker
        if best >= 0:
            cleaned = cleaned[best + len(best_marker):].strip()
    hard_stops = [
        "Amenities",
        "Olanaklar",
        "Hotel policies",
        "Otel politikaları",
        "Sign in",
        "Giriş yap",
        "Cookie",
    ]
    for marker in hard_stops:
        idx = cleaned.find(marker)
        if idx > 160:
            cleaned = cleaned[:idx]
    return cleaned.strip(" -|")


def looks_like_review_snippet(value: str) -> bool:
    lowered = value.lower()
    if len(value) < 80:
        return False
    if looks_blocked(value):
        return False
    if not any(keyword in lowered for keyword in REVIEW_KEYWORDS):
        return False
    boilerplate = ["check-in", "checkout", "chkin", "chkout", "terms and conditions"]
    return not any(term in lowered for term in boilerplate)


def infer_generic_rating(value: str) -> str:
    match = re.search(r"\b(?:rated\s*)?(\d+(?:\.\d+)?)\s*/\s*(?:5|10)\b", value, re.IGNORECASE)
    return match.group(0) if match else ""


def infer_generic_reviewer(value: str) -> str:
    compact = normalize_text(value)
    for marker in ("by ", "By ", "Guest ", "Misafir "):
        idx = compact.find(marker)
        if idx >= 0:
            return compact[idx : idx + 80].split(".")[0].strip()[:80]
    return ""


def dedupe_parsed_reviews(records: list[ParsedReview]) -> list[ParsedReview]:
    seen: set[tuple[str, str]] = set()
    unique: list[ParsedReview] = []
    for record in sorted(records, key=lambda item: item.date, reverse=True):
        key = (record.date.isoformat(), normalize_text(record.text)[:180])
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)
    return unique


def render_html_report(result: dict, *, result_path: str = DEFAULT_HTML) -> str:
    summary_rows = result["summary"]
    review_rows = result["reviews"]
    table_rows = []
    for row in summary_rows:
        table_rows.append(
            "<tr>"
            f"<td>{html.escape(row['hotel'])}</td>"
            f"<td>{html.escape(row['platform'])}</td>"
            f"<td>{html.escape(row['status'])}</td>"
            f"<td>{html.escape(row.get('fetch_status') or '')}</td>"
            f"<td class=\"num\">{html.escape(str(row['recent_review_count']))}</td>"
            f"<td>{html.escape(row.get('fetch_method') or '')}</td>"
            f"<td>{'yes' if row.get('blocked') else 'no'}</td>"
            f"<td>{html.escape(row.get('error') or '')}</td>"
            f"<td><a href=\"{html.escape(row['url'])}\">source</a></td>"
            "</tr>"
        )

    review_cards = []
    for review in review_rows:
        review_cards.append(
            f"""
            <article class="card">
              <h2>{html.escape(review['hotel'])} <span>{html.escape(review['platform'])}</span></h2>
              <div class="meta">{html.escape(review['date'])} {html.escape(review.get('rating') or '')}</div>
              <p>{html.escape(review.get('text') or '')}</p>
              <a href="{html.escape(review['source_url'])}">Review source</a>
            </article>
            """
        )

    xlsx_link = html.escape(relative_or_raw(DEFAULT_XLSX, result_path=result_path))
    json_link = html.escape(relative_or_raw(DEFAULT_JSON, result_path=result_path))
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>OTA Recent Reviews Report</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; background: #f7f8fb; color: #182033; }}
    header {{ background: #172033; color: white; padding: 28px 36px; }}
    h1 {{ margin: 0 0 8px; font-size: 28px; }}
    header p {{ margin: 0; color: #cbd5e1; }}
    main {{ max-width: 1240px; margin: 0 auto; padding: 24px; }}
    .metrics {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; margin-bottom: 18px; }}
    .metric, .table-wrap, .card {{ background: white; border: 1px solid #d9dee8; border-radius: 8px; }}
    .metric {{ padding: 16px; }}
    .metric b {{ display: block; font-size: 24px; }}
    .metric span {{ color: #647084; font-size: 13px; }}
    .links {{ margin: 0 0 18px; }}
    .table-wrap {{ overflow-x: auto; margin-bottom: 18px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 9px 11px; border-bottom: 1px solid #d9dee8; text-align: left; vertical-align: top; }}
    th {{ background: #eef2f7; font-size: 12px; color: #46546a; text-transform: uppercase; }}
    .num {{ text-align: right; }}
    .card {{ padding: 16px; margin: 12px 0; }}
    .card h2 {{ margin: 0 0 8px; font-size: 18px; }}
    .card h2 span {{ color: #2457d6; font-size: 14px; margin-left: 8px; }}
    .meta {{ color: #647084; font-weight: 700; margin-bottom: 8px; }}
    a {{ color: #2457d6; }}
    @media (max-width: 840px) {{ .metrics {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <header>
    <h1>OTA Recent Reviews Report</h1>
    <p>{html.escape(result['last_30_days_window']['start'])} to {html.escape(result['last_30_days_window']['end'])}</p>
  </header>
  <main>
    <section class="metrics">
      <div class="metric"><b>{result['total_links']}</b><span>Total links</span></div>
      <div class="metric"><b>{result['processed']}</b><span>Processed</span></div>
      <div class="metric"><b>{result['links_with_recent_reviews']}</b><span>Links with reviews</span></div>
      <div class="metric"><b>{result['total_recent_reviews']}</b><span>Recent reviews</span></div>
      <div class="metric"><b>{result.get('failed_or_blocked', result['blocked_or_empty'])}</b><span>Failed/blocked</span></div>
    </section>
    <p class="links"><a href="{xlsx_link}">Excel report</a> · <a href="{json_link}">JSON report</a></p>
    <section class="table-wrap">
      <table>
        <thead><tr><th>Hotel</th><th>Platform</th><th>Status</th><th>Fetch Status</th><th class="num">Recent Reviews</th><th>Method</th><th>Blocked</th><th>Error</th><th>Source</th></tr></thead>
        <tbody>{''.join(table_rows)}</tbody>
      </table>
    </section>
    {''.join(review_cards) if review_cards else '<p>No recent reviews were parsed in this run.</p>'}
  </main>
</body>
</html>"""


def write_results_workbook(output_path: Path, result: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Window Start", result["last_30_days_window"]["start"], "Window End", result["last_30_days_window"]["end"]])
    summary.append([])
    summary_headers = [
        "Hotel",
        "Platform",
        "Status",
        "Fetch Status",
        "Recent Review Count",
        "Fetch Method",
        "Blocked",
        "Text Length",
        "URL",
        "Final URL",
        "Error",
    ]
    summary.append(summary_headers)
    for row in result["summary"]:
        summary.append(
            [
                row.get("hotel"),
                row.get("platform"),
                row.get("status"),
                row.get("fetch_status"),
                row.get("recent_review_count"),
                row.get("fetch_method"),
                row.get("blocked"),
                row.get("text_length"),
                row.get("url"),
                row.get("final_url"),
                row.get("error"),
            ]
        )

    reviews = workbook.create_sheet("Reviews")
    review_headers = [
        "Hotel",
        "Platform",
        "Date",
        "Reviewer",
        "Rating",
        "Title",
        "Review Text",
        "Source URL",
        "Extraction Method",
    ]
    reviews.append(review_headers)
    for row in result["reviews"]:
        reviews.append(
            [
                row.get("hotel"),
                row.get("platform"),
                row.get("date"),
                row.get("reviewer"),
                row.get("rating"),
                row.get("title"),
                row.get("text"),
                row.get("source_url"),
                row.get("extraction_method"),
            ]
        )

    for worksheet in workbook.worksheets:
        header_row = 3 if worksheet.title == "Summary" else 1
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True)
        for column_cells in worksheet.columns:
            max_len = min(max(len(str(cell.value or "")) for cell in column_cells), 80)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, max_len + 2)
        worksheet.freeze_panes = "A4" if worksheet.title == "Summary" else "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)


def relative_or_raw(path: str, *, result_path: str) -> str:
    try:
        return os.path.relpath(Path(path).resolve(), Path(result_path).resolve().parent)
    except Exception:
        return path


def looks_blocked(text: str) -> bool:
    lowered = (text or "").lower()
    return any(marker in lowered for marker in BLOCKED_MARKERS)


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def sanitize_error(error: object) -> str:
    return normalize_text(str(error or ""))[:500]


def _as_of_date(configured: str | None = None) -> date:
    value = configured or os.getenv("OTA_REVIEW_AS_OF_DATE")
    if value:
        return date.fromisoformat(value)
    return date.today()
