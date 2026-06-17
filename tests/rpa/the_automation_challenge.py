"""The Automation Challenge workflow."""

from __future__ import annotations

import html
import json
import secrets
import string
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from harness import RPAWorkflow


TARGET_URL = "https://www.theautomationchallenge.com/"
DEFAULT_OUTPUT_DIR = "runs/the_automation_challenge"
DEFAULT_OUTPUT_HTML = "reports/the_automation_challenge/report_{timestamp}.html"
HEADERS = [
    "employer_identification_number",
    "company_name",
    "sector",
    "company_address",
    "automation_tool",
    "annual_automation_saving",
    "date_of_first_project",
]
FIELD_PREFIXES = [
    ("company_name", "company_name_input_field"),
    ("company_address", "address_input_field"),
    ("employer_identification_number", "ein_input_field"),
    ("sector", "sector_input_field"),
    ("automation_tool", "automation_tool_input_field"),
    ("annual_automation_saving", "annual_saving_input_field"),
    ("date_of_first_project", "date_input_field"),
]


class TheAutomationChallengeWorkflow(RPAWorkflow):
    name = "the_automation_challenge"
    tags = ["rpa", "browser", "excel", "external", "public-site"]
    max_retries_per_record = 0

    async def setup(self):
        variables = getattr(self.config, "variables", {}) or {}
        self.target_url = str(variables.get("automation_challenge_url") or TARGET_URL)
        self.run_label = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = (
            Path(variables.get("automation_challenge_output_dir") or DEFAULT_OUTPUT_DIR)
            / self.run_label
        )
        self.output_html = Path(
            str(variables.get("automation_challenge_output_html") or DEFAULT_OUTPUT_HTML).format(
                timestamp=self.run_label
            )
        )
        self.rows: list[dict[str, Any]] = []

    def get_records(self):
        yield {"id": f"the_automation_challenge_{self.run_label}"}

    async def process_record(self, record: dict) -> dict:
        run = await run_challenge(
            target_url=self.target_url,
            output_dir=self.output_dir,
            headless=bool(getattr(self.config, "headless", True)),
            browser_name=str(getattr(self.config, "browser", "chromium")),
            progress=lambda event, **fields: self._timeline(
                event,
                phase="Process Records",
                record_id=record["id"],
                **fields,
            ),
        )
        row = {
            "record_id": record["id"],
            **run,
            "side_effects": [
                "external_write: create throwaway challenge account",
                "external_read: download public challenge workbook",
                "external_write: submit 50 challenge rows",
            ],
            "next_action": "Archive the report and screenshots as challenge evidence."
            if run["status"] == "passed"
            else "Inspect final screenshot and result text before rerunning.",
        }
        self.rows.append(row)
        self.record_evidence(row, record=record, stage="the_automation_challenge")
        return {
            "status": run["status"],
            "reason": run.get("reason", ""),
            "details": row,
            "evidence_path": str(self.output_dir / "report.json"),
        }

    async def teardown(self):
        report_paths = write_report(
            rows=self.rows,
            output_json=self.output_dir / "report.json",
            output_html=self.output_html,
            metadata={
                "workflow": self.name,
                "target_url": getattr(self, "target_url", TARGET_URL),
                "started_at": getattr(self, "run_label", ""),
                "finished_at": datetime.now().isoformat(timespec="seconds"),
            },
        )
        self.result.output_files.extend([report_paths["json"], report_paths["html"]])
        for row in self.rows:
            self.result.screenshots.extend(row.get("screenshots", []))
        self.log(f"The Automation Challenge report JSON: {report_paths['json']}")
        self.log(f"The Automation Challenge report HTML: {report_paths['html']}")


async def run_challenge(
    *,
    target_url: str,
    output_dir: Path,
    headless: bool,
    browser_name: str,
    progress=None,
) -> dict[str, Any]:
    from playwright.async_api import async_playwright

    output_dir.mkdir(parents=True, exist_ok=True)
    screenshots: list[str] = []

    async with async_playwright() as playwright:
        browser_type = getattr(playwright, browser_name)
        browser = await browser_type.launch(headless=headless)
        page = await browser.new_page(viewport={"width": 1440, "height": 1000})
        await page.goto(target_url, wait_until="networkidle", timeout=60000)

        workbook_path = output_dir / "challenge.xlsx"
        async with page.expect_download(timeout=30000) as download_info:
            await page.get_by_text("Download Excel Spreadsheet").click(timeout=10000)
        await (await download_info.value).save_as(workbook_path)
        data = read_rows(workbook_path)
        if progress:
            progress(
                "challenge.workbook_downloaded",
                status="running",
                message=f"Workbook downloaded: {len(data)} rows",
            )

        before = output_dir / "before_start.png"
        await page.screenshot(path=str(before), full_page=True)
        screenshots.append(str(before))

        await sign_up_and_start(page)
        if progress:
            progress("challenge.started", status="running", message="Challenge started")
        started = output_dir / "after_start.png"
        await page.screenshot(path=str(started), full_page=True)
        screenshots.append(str(started))

        captcha_count = 0
        for index, row in enumerate(data, start=1):
            if progress:
                progress(
                    "challenge.row.started",
                    status="running",
                    message=f"Row {index}/{len(data)} started",
                )
            captcha_count += int(await clear_fake_recaptcha(page))
            await fill_row(page, row)
            await click_submit(page)
            captcha_count += int(await clear_fake_recaptcha(page))
            await wait_round(page, index, len(data))
            if progress:
                progress(
                    "challenge.row.submitted",
                    status="running",
                    message=f"Row {index}/{len(data)} submitted",
                )

        result_text = await page.locator("body").inner_text(timeout=10000)
        final = output_dir / "final.png"
        await page.screenshot(path=str(final), full_page=True)
        screenshots.append(str(final))
        await browser.close()

    passed = "SUCCESS!" in result_text and "100% (350 out of 350 fields)" in result_text
    (output_dir / "result.txt").write_text(result_text, encoding="utf-8")
    if progress:
        progress(
            "challenge.result",
            status="passed" if passed else "failed",
            message=result_text.splitlines()[0] if result_text.splitlines() else "",
        )
    return {
        "status": "passed" if passed else "failed",
        "reason": "" if passed else result_text[:500],
        "result_text": result_text,
        "workbook_path": str(workbook_path),
        "row_count": len(data),
        "captcha_cleared": captcha_count,
        "screenshots": screenshots,
    }


def read_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    if headers != HEADERS:
        raise ValueError(f"Unexpected workbook headers: {headers}")
    return [dict(zip(headers, map(str, values))) for values in sheet.iter_rows(min_row=2, values_only=True)]


def random_password() -> str:
    alphabet = string.ascii_letters + string.digits
    return "Rpa!" + "".join(secrets.choice(alphabet) for _ in range(16))


async def sign_up_and_start(page) -> None:
    await page.get_by_role("button", name="Start").click(timeout=10000)
    await page.wait_for_timeout(800)
    if await page.get_by_role("button", name="SIGN UP", exact=True).count():
        email = f"codex.rpa.{datetime.now().strftime('%Y%m%d%H%M%S')}@example.com"
        await page.get_by_placeholder("First Name").fill("Codex")
        await page.get_by_placeholder("Last Name").fill("RPA")
        await page.get_by_placeholder("Email").fill(email)
        await page.get_by_placeholder("Password").fill(random_password())
        await page.get_by_role("button", name="SIGN UP", exact=True).click(timeout=10000)
        await page.wait_for_timeout(2500)
    await page.get_by_role("button", name="Start").click(timeout=10000)
    await page.wait_for_timeout(500)


async def clear_fake_recaptcha(page) -> bool:
    popup = page.locator(".bubble-element.Popup:visible").filter(has_text="reCAPTCHA")
    if not await popup.count():
        return False
    await popup.locator("button:visible").first.click(timeout=5000)
    await popup.wait_for(state="hidden", timeout=10000)
    return True


async def fill_row(page, row: dict[str, str]) -> None:
    for column, prefix in FIELD_PREFIXES:
        await clear_fake_recaptcha(page)
        await page.locator(f'input[id^="{prefix}"]:visible').first.fill(row[column], timeout=10000)


async def click_submit(page) -> None:
    for _ in range(3):
        await clear_fake_recaptcha(page)
        try:
            await page.get_by_role("button", name="Submit").click(timeout=5000)
            return
        except Exception:
            if not await clear_fake_recaptcha(page):
                raise
    await page.get_by_role("button", name="Submit").click(timeout=5000)


async def wait_round(page, index: int, total: int) -> None:
    if index >= total:
        await page.wait_for_timeout(1200)
        return
    target = f"Round {index + 1} of {total}"
    for _ in range(30):
        await clear_fake_recaptcha(page)
        if target in await page.locator("body").inner_text(timeout=2000):
            return
        await page.wait_for_timeout(100)
    raise TimeoutError(f"Round did not advance to {target}")


def write_report(
    *,
    rows: list[dict[str, Any]],
    output_json: Path,
    output_html: Path,
    metadata: dict[str, Any],
) -> dict[str, str]:
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_html.parent.mkdir(parents=True, exist_ok=True)
    passed = sum(1 for row in rows if row.get("status") == "passed")
    report = {
        "metadata": metadata,
        "summary": {
            "total": len(rows),
            "passed": passed,
            "failed": len(rows) - passed,
            "status": "passed" if rows and passed == len(rows) else "failed",
            "next_action": first_value(rows, "next_action"),
        },
        "records": rows,
    }
    output_json.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    output_html.write_text(render_html(report), encoding="utf-8")
    return {"json": str(output_json), "html": str(output_html)}


def render_html(report: dict[str, Any]) -> str:
    cards = []
    for row in report.get("records", []):
        screenshots = "".join(
            f"<li><code>{html.escape(str(path))}</code></li>" for path in row.get("screenshots", [])
        )
        cards.append(
            f"<section><h2>{html.escape(str(row.get('record_id', 'record')))}</h2>"
            f"<p><strong>Status:</strong> {html.escape(str(row.get('status', '')))}</p>"
            f"<p><strong>Rows:</strong> {row.get('row_count', 0)}</p>"
            f"<p><strong>Fake CAPTCHA cleared:</strong> {row.get('captcha_cleared', 0)}</p>"
            f"<pre>{html.escape(str(row.get('result_text', row.get('reason', ''))))}</pre>"
            f"<h3>Screenshots</h3><ul>{screenshots}</ul></section>"
        )
    summary = report.get("summary", {})
    return (
        "<!doctype html><html><head><meta charset='utf-8'><title>The Automation Challenge</title>"
        "<style>body{font-family:Segoe UI,Arial,sans-serif;margin:32px;background:#f6f7f9;color:#1f2937}"
        "main{max-width:1100px;margin:auto;background:white;padding:28px;border:1px solid #d1d5db}"
        "section{border-top:1px solid #e5e7eb;margin-top:24px;padding-top:16px}"
        "pre{white-space:pre-wrap;background:#f3f4f6;padding:12px}</style></head><body><main>"
        "<h1>The Automation Challenge</h1>"
        f"<p><strong>Status:</strong> {html.escape(str(summary.get('status', 'unknown')))}</p>"
        f"<p><strong>Next action:</strong> {html.escape(str(summary.get('next_action', '')))}</p>"
        f"{''.join(cards)}</main></body></html>"
    )


def first_value(rows: list[dict[str, Any]], key: str) -> str:
    for row in rows:
        if row.get(key):
            return str(row[key])
    return ""
