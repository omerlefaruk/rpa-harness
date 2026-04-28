"""Capability characterization for Python RPAWorkflow and Excel workflows."""

from pathlib import Path
from typing import Any

import openpyxl
import pytest

from harness.config import HarnessConfig
from harness.rpa.excel import ExcelHandler
from harness.rpa.workflow import RPAWorkflow, WorkflowStatus
from harness.verification import CheckRunner, CheckType, SuccessCheck


def _config(tmp_path: Path, **variables: Any) -> HarnessConfig:
    return HarnessConfig(
        headless=True,
        enable_vision=False,
        report_dir=str(tmp_path / "reports"),
        screenshot_dir=str(tmp_path / "screenshots"),
        variables=variables,
    )


async def _execute(workflow: RPAWorkflow):
    return await workflow._execute()


class ZeroRecordsWorkflow(RPAWorkflow):
    name = "zero_records_capability"
    tags = ["rpa", "capability"]

    def get_records(self):
        return iter([])

    async def process_record(self, record):
        raise AssertionError("zero-record workflow should not process records")


@pytest.mark.asyncio
async def test_zero_records_workflow_passes_with_no_processed_records(tmp_path):
    result = await _execute(ZeroRecordsWorkflow(_config(tmp_path)))

    assert result.status == WorkflowStatus.PASSED
    assert result.total_records == 0
    assert result.processed_records == 0
    assert result.failed_records == 0


class AllPassWorkflow(RPAWorkflow):
    name = "all_pass_capability"
    tags = ["rpa", "capability"]

    def __init__(self, config):
        super().__init__(config)
        self.successes: list[str] = []

    def get_records(self):
        return iter([{"id": "A"}, {"id": "B"}])

    async def process_record(self, record):
        return {"status": "passed", "record_id": record["id"]}

    async def on_success(self, record, details=None):
        self.successes.append(record["id"])


@pytest.mark.asyncio
async def test_all_records_pass_and_on_success_runs(tmp_path):
    workflow = AllPassWorkflow(_config(tmp_path))

    result = await _execute(workflow)

    assert result.status == WorkflowStatus.PASSED
    assert result.total_records == 2
    assert result.processed_records == 2
    assert result.failed_records == 0
    assert workflow.successes == ["A", "B"]


class PassMismatchWorkflow(RPAWorkflow):
    name = "pass_mismatch_capability"
    tags = ["rpa", "excel", "capability"]
    allow_mismatches = True

    def __init__(self, config):
        super().__init__(config)
        self.output_path = Path(config.variables["output_file"])
        self.output_excel = None

    async def setup(self):
        self.output_excel = ExcelHandler(str(self.output_path))
        self.output_excel.write_rows(
            sheet="Mismatches",
            headers=["ID", "Reason", "Expected", "Actual"],
        )

    def get_records(self):
        return iter(
            [
                {"id": "PASS", "expected": "OK", "actual": "OK"},
                {"id": "MISS", "expected": "OK", "actual": "BAD"},
            ]
        )

    async def process_record(self, record):
        if record["expected"] == record["actual"]:
            return {"status": "passed"}
        return {
            "status": "mismatch",
            "reason": "Value mismatch",
            "details": {"expected": record["expected"], "actual": record["actual"]},
        }

    async def on_mismatch(self, record, reason, details=None):
        self.output_excel.append_row(
            sheet="Mismatches",
            mapping={
                "ID": record["id"],
                "Reason": reason,
                "Expected": details["expected"],
                "Actual": details["actual"],
            },
            headers=["ID", "Reason", "Expected", "Actual"],
        )

    async def teardown(self):
        self.output_excel.save()
        self.output_excel.close()
        self.result.output_files.append(str(self.output_path))


@pytest.mark.asyncio
async def test_one_pass_one_mismatch_can_be_explicitly_allowed_and_writes_output(tmp_path):
    output_file = tmp_path / "mismatches.xlsx"
    workflow = PassMismatchWorkflow(_config(tmp_path, output_file=str(output_file)))

    result = await _execute(workflow)

    assert result.status == WorkflowStatus.PASSED
    assert result.processed_records == 1
    assert result.failed_records == 1
    assert result.output_files == [str(output_file)]

    workbook = openpyxl.load_workbook(output_file, data_only=True)
    try:
        sheet = workbook["Mismatches"]
        assert sheet["A2"].value == "MISS"
        assert sheet["B2"].value == "Value mismatch"
        assert sheet["C2"].value == "OK"
        assert sheet["D2"].value == "BAD"
    finally:
        workbook.close()


class StrictPassMismatchWorkflow(PassMismatchWorkflow):
    name = "strict_pass_mismatch_capability"
    allow_mismatches = False


@pytest.mark.asyncio
async def test_one_pass_one_mismatch_fails_by_default_without_allow_mismatches(tmp_path):
    output_file = tmp_path / "strict_mismatches.xlsx"
    workflow = StrictPassMismatchWorkflow(_config(tmp_path, output_file=str(output_file)))

    result = await _execute(workflow)

    assert result.status == WorkflowStatus.FAILED
    assert result.processed_records == 1
    assert result.failed_records == 1
    assert result.output_files == [str(output_file)]


class SkippedRecordWorkflow(RPAWorkflow):
    name = "skipped_record_capability"
    tags = ["rpa", "capability"]

    def get_records(self):
        return iter([{"id": "SKIP"}])

    async def process_record(self, record):
        return {"status": "skipped", "reason": "No action required"}


@pytest.mark.asyncio
async def test_skipped_record_is_counted_without_failure(tmp_path):
    result = await _execute(SkippedRecordWorkflow(_config(tmp_path)))

    assert result.status == WorkflowStatus.PASSED
    assert result.skipped_records == 1
    assert result.failed_records == 0


class RetryWorkflow(RPAWorkflow):
    name = "retry_record_capability"
    tags = ["rpa", "capability"]
    max_retries_per_record = 1

    def __init__(self, config, always_retry: bool = False):
        super().__init__(config)
        self.always_retry = always_retry
        self.calls = 0

    def get_records(self):
        return iter([{"id": "RETRY"}])

    async def process_record(self, record):
        self.calls += 1
        if self.always_retry or self.calls == 1:
            return {"status": "retry", "reason": "temporary"}
        return {"status": "passed"}


@pytest.mark.asyncio
async def test_retryable_record_succeeds_on_second_attempt(tmp_path):
    workflow = RetryWorkflow(_config(tmp_path))

    result = await _execute(workflow)

    assert result.status == WorkflowStatus.PASSED
    assert result.processed_records == 1
    assert result.retried_records == 1
    assert workflow.calls == 2


@pytest.mark.asyncio
async def test_retryable_record_exhausts_attempts_and_fails_when_no_record_passes(tmp_path):
    workflow = RetryWorkflow(_config(tmp_path), always_retry=True)

    result = await _execute(workflow)

    assert result.status == WorkflowStatus.FAILED
    assert result.processed_records == 0
    assert result.failed_records == 1
    assert result.retried_records == 1
    assert workflow.calls == 2


def _create_input_workbook(path: Path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Input"
    sheet.append(["ID", "Expected", "Actual", "Optional"])
    sheet.append(["PASS", "A", "A", None])
    sheet.append(["MISS", "A", "B", ""])
    workbook.save(path)
    workbook.close()


def _normalize(value):
    if value is None:
        return ""
    return str(value).strip()


class ExcelComparisonWorkflow(RPAWorkflow):
    name = "excel_comparison_capability"
    tags = ["rpa", "excel", "capability"]
    allow_mismatches = True

    def __init__(self, config):
        super().__init__(config)
        self.input_path = Path(config.variables["input_file"])
        self.output_path = Path(config.variables["output_file"])

    async def setup(self):
        if not self.input_path.exists():
            raise FileNotFoundError(f"Input workbook not found: {self.input_path}")
        self.input_excel = ExcelHandler(str(self.input_path))
        self.output_excel = ExcelHandler(str(self.output_path))
        self.output_excel.write_rows(
            sheet="Mismatches",
            headers=["ID", "Expected", "Actual", "Reason"],
        )

    def get_records(self):
        for row in self.input_excel.iter_rows(sheet="Input"):
            yield {
                "id": _normalize(row.get("ID")),
                "expected": _normalize(row.get("Expected")),
                "actual": _normalize(row.get("Actual")),
                "optional": _normalize(row.get("Optional")),
            }

    async def process_record(self, record):
        if record["expected"] == record["actual"]:
            return {"status": "passed"}
        return {
            "status": "mismatch",
            "reason": "Expected and actual differ",
            "details": {"expected": record["expected"], "actual": record["actual"]},
        }

    async def on_mismatch(self, record, reason, details=None):
        self.output_excel.append_row(
            sheet="Mismatches",
            mapping={
                "ID": record["id"],
                "Expected": details["expected"],
                "Actual": details["actual"],
                "Reason": reason,
            },
            headers=["ID", "Expected", "Actual", "Reason"],
        )

    async def teardown(self):
        if getattr(self, "input_excel", None):
            self.input_excel.close()
        if getattr(self, "output_excel", None):
            self.output_excel.save()
            self.output_excel.close()
            self.result.output_files.append(str(self.output_path))


@pytest.mark.asyncio
async def test_excel_data_driven_workflow_reads_normalizes_compares_and_verifies_output(tmp_path):
    input_file = tmp_path / "input.xlsx"
    output_file = tmp_path / "mismatches.xlsx"
    _create_input_workbook(input_file)

    workflow = ExcelComparisonWorkflow(
        _config(tmp_path, input_file=str(input_file), output_file=str(output_file))
    )
    result = await _execute(workflow)

    assert result.status == WorkflowStatus.PASSED
    assert result.total_records == 2
    assert result.processed_records == 1
    assert result.failed_records == 1
    assert result.output_files == [str(output_file)]

    runner = CheckRunner()
    runner.set_context("workbook_path", str(output_file))
    runner.set_context("sheet_name", "Mismatches")
    checks = [
        SuccessCheck(type=CheckType.WORKBOOK_EXISTS, value=str(output_file)),
        SuccessCheck(type=CheckType.SHEET_EXISTS, value="Mismatches"),
        SuccessCheck(
            type=CheckType.CELL_EQUALS,
            value={"sheet": "Mismatches", "cell": "A2", "value": "MISS"},
        ),
    ]

    assert all(runner.run(check).passed for check in checks)


@pytest.mark.asyncio
async def test_missing_input_workbook_gives_predictable_workflow_failure(tmp_path):
    missing_input = tmp_path / "missing.xlsx"
    output_file = tmp_path / "mismatches.xlsx"

    result = await _execute(
        ExcelComparisonWorkflow(
            _config(tmp_path, input_file=str(missing_input), output_file=str(output_file))
        )
    )

    assert result.status == WorkflowStatus.FAILED
    assert "Input workbook not found" in result.error_message


def test_excel_handler_can_reject_missing_input_without_creating_workbook(tmp_path):
    missing_input = tmp_path / "missing.xlsx"

    with pytest.raises(FileNotFoundError, match="Workbook not found"):
        ExcelHandler(str(missing_input), create_if_missing=False)

    assert not missing_input.exists()
