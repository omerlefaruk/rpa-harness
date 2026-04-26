"""
Excel handler for RPA data-driven workflows.
Adapted from automation-harness with added validation and CSV fallback.
"""

from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Union
from dataclasses import dataclass, field

from harness.logger import HarnessLogger


@dataclass
class ExcelRow:
    row_number: int
    data: Dict[str, Any] = field(default_factory=dict)
    raw_values: List[Any] = field(default_factory=list)

    def get(self, column: str, default=None) -> Any:
        return self.data.get(column, default)

    def get_by_index(self, index: int, default=None) -> Any:
        if 0 <= index < len(self.raw_values):
            return self.raw_values[index]
        return default

    def is_empty(self) -> bool:
        return all(v is None or str(v).strip() == "" for v in self.raw_values)

    def __repr__(self) -> str:
        return f"ExcelRow({self.row_number}, {self.data})"


class ExcelHandler:
    def __init__(self, file_path: str, logger: Optional[HarnessLogger] = None):
        self.file_path = Path(file_path)
        self.logger = logger or HarnessLogger("excel")
        self._workbook = None
        self._open_workbook()

    def _open_workbook(self):
        try:
            import openpyxl
            if self.file_path.exists():
                self._workbook = openpyxl.load_workbook(str(self.file_path), data_only=True)
                self.logger.info(f"Opened workbook: {self.file_path}")
            else:
                self._workbook = openpyxl.Workbook()
                self.logger.info(f"Created new workbook: {self.file_path}")
        except ImportError:
            self.logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise

    def iter_rows(
        self,
        sheet: str = None,
        header_row: int = 1,
        min_row: int = None,
        max_row: int = None,
        columns: List[str] = None,
    ) -> Iterator[ExcelRow]:
        ws = self._get_sheet(sheet)

        headers = []
        if header_row and header_row > 0:
            headers = [
                str(cell.value) if cell.value else f"Col_{i}"
                for i, cell in enumerate(ws[header_row], 1)
            ]

        if columns:
            headers = columns

        start = min_row if min_row else (header_row + 1 if header_row else 1)
        end = max_row if max_row else ws.max_row

        for row_idx in range(start, end + 1):
            row_data = ws[row_idx]
            raw = [cell.value for cell in row_data]

            if all(v is None or str(v).strip() == "" for v in raw):
                continue

            data = {}
            for i, header in enumerate(headers):
                if i < len(raw):
                    data[header] = raw[i]

            yield ExcelRow(row_number=row_idx, data=data, raw_values=raw)

    def read_column(
        self,
        sheet: str = None,
        column: Union[str, int] = "A",
        header_row: int = 1,
        skip_header: bool = True,
        data_type: str = "auto",
    ) -> List[Any]:
        ws = self._get_sheet(sheet)

        if isinstance(column, str):
            col_idx = self._col_letter_to_index(column)
        else:
            col_idx = column

        values = []
        start_row = header_row + 1 if skip_header and header_row else 1

        for row in ws.iter_rows(
            min_row=start_row, max_row=ws.max_row,
            min_col=col_idx, max_col=col_idx, values_only=True,
        ):
            val = row[0]
            if val is None:
                continue
            values.append(self._convert_type(val, data_type))

        return values

    def read_cell(self, sheet: str = None, cell: str = "A1", data_type: str = "auto") -> Any:
        ws = self._get_sheet(sheet)
        val = ws[cell].value
        return self._convert_type(val, data_type)

    def write_cell(self, sheet: str = None, cell: str = "A1", value: Any = None):
        ws = self._get_sheet(sheet)
        ws[cell] = value

    def write_rows(
        self,
        sheet: str = None,
        headers: List[str] = None,
        rows: List[List[Any]] = None,
        start_row: int = 1,
    ):
        ws = self._get_or_create_sheet(sheet)
        current_row = start_row

        if headers:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col_idx, value=header)
            current_row += 1

        if rows:
            for row_data in rows:
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                current_row += 1

        self.logger.info(f"Wrote {len(rows or [])} rows to {ws.title}")

    def append_row(
        self,
        sheet: str = None,
        row_data: List[Any] = None,
        mapping: Dict[str, Any] = None,
        headers: List[str] = None,
    ):
        ws = self._get_or_create_sheet(sheet)

        if mapping and headers:
            row_data = [mapping.get(h, "") for h in headers]

        if row_data:
            ws.append(row_data)

    def validate_columns(self, sheet: str = None, expected_columns: List[str] = None) -> bool:
        ws = self._get_sheet(sheet)
        header_row = ws[1]
        actual = [str(cell.value) for cell in header_row if cell.value]

        if expected_columns:
            missing = [c for c in expected_columns if c not in actual]
            if missing:
                self.logger.warning(f"Missing columns: {missing}")
                return False
        return True

    def create_sheet(self, name: str):
        if name not in self._workbook.sheetnames:
            self._workbook.create_sheet(title=name)
            self.logger.info(f"Created sheet: {name}")

    def save(self, path: str = None):
        save_path = Path(path) if path else self.file_path
        save_path.parent.mkdir(parents=True, exist_ok=True)
        self._workbook.save(str(save_path))
        self.logger.info(f"Saved workbook: {save_path}")

    def sheet_names(self) -> List[str]:
        return self._workbook.sheetnames

    def row_count(self, sheet: str = None, skip_header: bool = True) -> int:
        ws = self._get_sheet(sheet)
        total = ws.max_row
        return total - 1 if skip_header and total > 1 else total

    def close(self):
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def to_csv(self, sheet: str = None, output_path: str = None) -> str:
        import csv
        ws = self._get_sheet(sheet)
        csv_path = output_path or str(self.file_path.with_suffix(".csv"))

        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        self.logger.info(f"Exported to CSV: {csv_path}")
        return csv_path

    def _get_sheet(self, name: str = None):
        if name and name in self._workbook.sheetnames:
            return self._workbook[name]
        return self._workbook.active

    def _get_or_create_sheet(self, name: str = None):
        if name:
            if name not in self._workbook.sheetnames:
                self._workbook.create_sheet(title=name)
            return self._workbook[name]
        return self._workbook.active

    @staticmethod
    def _col_letter_to_index(col: str) -> int:
        result = 0
        for char in col.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result

    @staticmethod
    def _convert_type(value: Any, data_type: str) -> Any:
        if value is None:
            return None

        if data_type == "auto":
            return value
        elif data_type == "str":
            return str(value)
        elif data_type == "int":
            try:
                return int(value)
            except (ValueError, TypeError):
                return value
        elif data_type == "float":
            try:
                return float(value)
            except (ValueError, TypeError):
                return value
        elif data_type == "bool":
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                return value.lower() in ("true", "yes", "1", "evet")
            return bool(value)
        return value
