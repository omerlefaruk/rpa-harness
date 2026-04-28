"""
Verification checks — executes success checks against real state.
"""
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from harness.verification.contract import CheckType, SuccessCheck, VerificationResult


class CheckRunner:
    def __init__(self):
        self._context: Dict[str, Any] = {}

    def set_context(self, key: str, value: Any):
        self._context[key] = value

    def run(self, check: SuccessCheck) -> VerificationResult:
        handler = {
            CheckType.FILE_EXISTS: self._check_file_exists,
            CheckType.VARIABLE_HAS_VALUE: self._check_variable_has_value,
            CheckType.VARIABLE_EQUALS: self._check_variable_equals,
            CheckType.TEXT_CONTAINS: self._check_text_contains,
            CheckType.ALWAYS_PASS: self._check_always_pass,
            CheckType.URL_CONTAINS: self._check_url_contains,
            CheckType.URL_EQUALS: self._check_url_equals,
            CheckType.VISIBLE_TEXT: self._check_visible_text,
            CheckType.SELECTOR_VISIBLE: self._check_selector_visible,
            CheckType.SELECTOR_HIDDEN: self._check_selector_hidden,
            CheckType.FIELD_HAS_VALUE: self._check_field_has_value,
            CheckType.DOWNLOAD_EXISTS: self._check_download_exists,
            CheckType.STATUS_CODE: self._check_status_code,
            CheckType.JSON_PATH_EQUALS: self._check_json_path_equals,
            CheckType.RESPONSE_CONTAINS: self._check_response_contains,
            CheckType.WORKBOOK_EXISTS: self._check_file_exists,
            CheckType.SHEET_EXISTS: self._check_sheet_exists,
            CheckType.CELL_EQUALS: self._check_cell_equals,
            CheckType.WINDOW_EXISTS: self._check_window_exists,
            CheckType.ELEMENT_EXISTS: self._check_element_exists,
            CheckType.ELEMENT_TEXT_EQUALS: self._check_element_text_equals,
        }

        fn = handler.get(check.type, self._check_unknown)
        try:
            return fn(check)
        except Exception as e:
            return VerificationResult(
                passed=False,
                check_type=check.type,
                expected=check.value,
                actual=f"Error: {e}",
                message=str(e),
            )

    def _check_file_exists(self, check: SuccessCheck) -> VerificationResult:
        path = Path(str(check.value))
        exists = path.exists()
        return VerificationResult(
            passed=exists,
            check_type=check.type,
            expected=str(check.value),
            actual=f"exists={exists}",
            message="" if exists else f"File not found: {check.value}",
        )

    def _check_download_exists(self, check: SuccessCheck) -> VerificationResult:
        expected = str(check.value)
        expected_path = Path(expected)

        candidates = self._collect_paths(
            "downloaded_file",
            "downloaded_files",
            "download_path",
            "download_paths",
            "output_files",
            "file_path",
        )

        if expected_path.exists():
            actual = str(expected_path)
            return VerificationResult(
                passed=True,
                check_type=check.type,
                expected=expected,
                actual=actual,
            )

        match = next(
            (
                candidate for candidate in candidates
                if candidate.exists()
                and (
                    str(candidate) == expected
                    or candidate.name == expected_path.name
                )
            ),
            None,
        )

        return VerificationResult(
            passed=match is not None,
            check_type=check.type,
            expected=expected,
            actual=str(match) if match else [str(candidate) for candidate in candidates],
            message="" if match else f"Download not found: {expected}",
        )

    def _check_variable_has_value(self, check: SuccessCheck) -> VerificationResult:
        value = self._context.get(str(check.value))
        has = value is not None and value != ""
        return VerificationResult(
            passed=has,
            check_type=check.type,
            expected=f"non-empty value",
            actual=str(value)[:100] if value else "None/empty",
            message="" if has else f"Variable '{check.value}' has no value",
        )

    def _check_variable_equals(self, check: SuccessCheck) -> VerificationResult:
        expected = check.value.get("value") if isinstance(check.value, dict) else check.value
        var = check.value.get("var") if isinstance(check.value, dict) else check.value
        actual = self._context.get(var)
        match = str(actual) == str(expected)
        return VerificationResult(
            passed=match,
            check_type=check.type,
            expected=expected,
            actual=actual,
            message="" if match else f"Expected '{expected}', got '{actual}'",
        )

    def _check_text_contains(self, check: SuccessCheck) -> VerificationResult:
        text = self._context.get("last_text", "")
        contains = str(check.value) in text
        return VerificationResult(
            passed=contains,
            check_type=check.type,
            expected=f"text contains '{check.value}'",
            actual=text[:200],
            message="" if contains else f"Text did not contain '{check.value}'",
        )

    def _check_always_pass(self, check: SuccessCheck) -> VerificationResult:
        return VerificationResult(
            passed=True, check_type=check.type,
            expected="always", actual="passed", message="No-op check",
        )

    def _check_url_contains(self, check: SuccessCheck) -> VerificationResult:
        url = self._context.get("current_url", "")
        contains = str(check.value) in url
        return VerificationResult(
            passed=contains,
            check_type=check.type,
            expected=check.value,
            actual=url[:200],
            message="" if contains else f"URL does not contain '{check.value}'",
        )

    def _check_url_equals(self, check: SuccessCheck) -> VerificationResult:
        url = self._context.get("current_url", "")
        match = url == str(check.value)
        return VerificationResult(
            passed=match, check_type=check.type,
            expected=check.value, actual=url,
            message="" if match else f"URL mismatch",
        )

    def _check_visible_text(self, check: SuccessCheck) -> VerificationResult:
        visible = self._context.get("visible_text", "")
        contains = str(check.value) in visible
        return VerificationResult(
            passed=contains,
            check_type=check.type,
            expected=check.value,
            actual=visible[:200],
            message="" if contains else f"Text not visible: '{check.value}'",
        )

    def _check_selector_visible(self, check: SuccessCheck) -> VerificationResult:
        visible = self._context.get("selector_visible", False)
        return VerificationResult(
            passed=visible, check_type=check.type,
            expected="element visible", actual=str(visible),
        )

    def _check_selector_hidden(self, check: SuccessCheck) -> VerificationResult:
        hidden = self._context.get("selector_hidden", True)
        return VerificationResult(
            passed=hidden, check_type=check.type,
            expected="element hidden", actual=str(hidden),
        )

    def _check_field_has_value(self, check: SuccessCheck) -> VerificationResult:
        value = self._context.get("field_value", "")
        has = bool(value)
        evidence = {}
        if check.redacted:
            evidence["redacted"] = True
            return VerificationResult(
                passed=has, check_type=check.type,
                expected="[REDACTED]", actual="[REDACTED]",
                message="Field has value (value redacted)", evidence=evidence,
            )
        return VerificationResult(
            passed=has, check_type=check.type,
            expected="non-empty", actual=str(value)[:100],
        )

    def _check_status_code(self, check: SuccessCheck) -> VerificationResult:
        code = self._context.get("status_code", 0)
        match = code == int(check.value)
        return VerificationResult(
            passed=match, check_type=check.type,
            expected=check.value, actual=str(code),
        )

    def _check_json_path_equals(self, check: SuccessCheck) -> VerificationResult:
        payload = check.value if isinstance(check.value, dict) else {}
        path = payload.get("path")
        expected = payload.get("value")
        response_data = self._context.get("response_json")

        if response_data is None:
            response_body = self._context.get("response_body")
            if isinstance(response_body, str) and response_body:
                try:
                    response_data = json.loads(response_body)
                except json.JSONDecodeError:
                    response_data = None

        if not path:
            return VerificationResult(
                passed=False,
                check_type=check.type,
                expected=expected,
                actual=None,
                message="json_path_equals requires a 'path' value",
            )

        found, actual = self._resolve_json_path(response_data, str(path))
        passed = found and self._values_equal(actual, expected)
        return VerificationResult(
            passed=passed,
            check_type=check.type,
            expected=expected,
            actual=actual,
            message="" if passed else f"JSON path '{path}' expected '{expected}', got '{actual}'",
        )

    def _check_response_contains(self, check: SuccessCheck) -> VerificationResult:
        body = self._context.get("response_body", "")
        contains = str(check.value) in body
        return VerificationResult(
            passed=contains, check_type=check.type,
            expected=f"contains '{check.value}'", actual=body[:500],
        )

    def _check_sheet_exists(self, check: SuccessCheck) -> VerificationResult:
        expected_sheet = str(check.value)
        sheet_names = self._get_sheet_names()
        exists = expected_sheet in sheet_names
        return VerificationResult(
            passed=exists,
            check_type=check.type,
            expected=expected_sheet,
            actual=sheet_names,
            message="" if exists else f"Sheet not found: {expected_sheet}",
        )

    def _check_cell_equals(self, check: SuccessCheck) -> VerificationResult:
        payload = check.value if isinstance(check.value, dict) else {}
        cell_ref = payload.get("cell")
        expected = payload.get("value")
        sheet_name = payload.get("sheet") or self._context.get("sheet_name")

        actual = self._get_cell_value(cell_ref, sheet_name)
        passed = str(actual) == str(expected)
        return VerificationResult(
            passed=passed,
            check_type=check.type,
            expected=expected,
            actual=actual,
            message="" if passed else f"Cell {cell_ref} expected '{expected}', got '{actual}'",
            evidence={"cell": cell_ref, "sheet": sheet_name} if cell_ref else {},
        )

    def _check_window_exists(self, check: SuccessCheck) -> VerificationResult:
        expected = str(check.value)
        explicit = self._context.get("window_exists")
        if isinstance(explicit, bool):
            passed = explicit
            actual = self._context.get("window_title") or self._context.get("current_window")
        else:
            windows = self._context.get("available_windows") or []
            actual = self._context.get("window_title") or self._context.get("current_window") or windows
            passed = False
            if isinstance(actual, str):
                passed = expected.lower() in actual.lower()
            elif isinstance(windows, list):
                passed = any(expected.lower() in str(window).lower() for window in windows)

        return VerificationResult(
            passed=passed,
            check_type=check.type,
            expected=expected,
            actual=actual,
            message="" if passed else f"Window not found: {expected}",
        )

    def _check_element_exists(self, check: SuccessCheck) -> VerificationResult:
        selector = check.selector or (check.value if isinstance(check.value, dict) else None)
        explicit = self._context.get("element_exists")
        actual = explicit

        if isinstance(explicit, bool):
            passed = explicit
        else:
            candidates = self._context.get("elements") or []
            actual = candidates
            passed = self._match_selector(selector, candidates)
            if not passed and "selector_visible" in self._context:
                passed = bool(self._context.get("selector_visible"))
                actual = self._context.get("selector_visible")

        return VerificationResult(
            passed=passed,
            check_type=check.type,
            expected=selector or check.value,
            actual=actual,
            message="" if passed else f"Element not found: {selector or check.value}",
        )

    def _check_element_text_equals(self, check: SuccessCheck) -> VerificationResult:
        expected = check.value.get("value") if isinstance(check.value, dict) else check.value
        actual = (
            self._context.get("element_text")
            or self._context.get("text")
            or self._context.get("last_text")
        )
        passed = str(actual) == str(expected)
        return VerificationResult(
            passed=passed,
            check_type=check.type,
            expected=expected,
            actual=actual,
            message="" if passed else f"Expected element text '{expected}', got '{actual}'",
        )

    def _check_unknown(self, check: SuccessCheck) -> VerificationResult:
        return VerificationResult(
            passed=False, check_type=check.type,
            expected=check.value, actual="unknown check type",
            message=f"No handler for check type: {check.type.value}",
        )

    def _collect_paths(self, *keys: str) -> List[Path]:
        paths: List[Path] = []
        for key in keys:
            value = self._context.get(key)
            if isinstance(value, (str, Path)):
                paths.append(Path(value))
            elif isinstance(value, list):
                paths.extend(Path(item) for item in value if isinstance(item, (str, Path)))
        return paths

    def _get_sheet_names(self) -> List[str]:
        if isinstance(self._context.get("sheet_names"), list):
            return [str(name) for name in self._context["sheet_names"]]

        workbook = self._context.get("workbook")
        if workbook is not None and hasattr(workbook, "sheetnames"):
            return [str(name) for name in workbook.sheetnames]

        workbook_path = self._context.get("workbook_path") or self._context.get("file_path")
        if workbook_path:
            try:
                import openpyxl

                loaded = openpyxl.load_workbook(str(workbook_path), data_only=True)
                try:
                    return [str(name) for name in loaded.sheetnames]
                finally:
                    loaded.close()
            except Exception:
                return []

        return []

    def _get_cell_value(self, cell_ref: Optional[str], sheet_name: Optional[str]) -> Any:
        if not cell_ref:
            return None

        cell_values = self._context.get("cell_values")
        if isinstance(cell_values, dict):
            if sheet_name and f"{sheet_name}!{cell_ref}" in cell_values:
                return cell_values.get(f"{sheet_name}!{cell_ref}")
            if cell_ref in cell_values:
                return cell_values.get(cell_ref)

        workbook = self._context.get("workbook")
        if workbook is not None:
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
            return worksheet[cell_ref].value

        workbook_path = self._context.get("workbook_path") or self._context.get("file_path")
        if workbook_path:
            try:
                import openpyxl

                loaded = openpyxl.load_workbook(str(workbook_path), data_only=True)
                try:
                    worksheet = loaded[sheet_name] if sheet_name and sheet_name in loaded.sheetnames else loaded.active
                    return worksheet[cell_ref].value
                finally:
                    loaded.close()
            except Exception:
                return None

        return None

    def _match_selector(self, selector: Any, candidates: List[Any]) -> bool:
        if not isinstance(selector, dict) or not candidates:
            return False

        strategy = str(selector.get("strategy", "")).lower()
        expected = selector.get("value")
        role = selector.get("role")
        name = selector.get("name")

        for candidate in candidates:
            if not isinstance(candidate, dict):
                continue
            if strategy == "automation_id" and str(candidate.get("automation_id", "")) == str(expected):
                return True
            if strategy == "name" and str(candidate.get("name", "")) == str(expected):
                return True
            if strategy == "id" and str(candidate.get("id", "")) == str(expected):
                return True
            if strategy == "text" and str(candidate.get("text", "")) == str(expected):
                return True
            if strategy == "role" and role and name:
                if str(candidate.get("role", "")) == str(role) and str(candidate.get("name", "")) == str(name):
                    return True
        return False

    def _resolve_json_path(self, data: Any, path: str) -> tuple[bool, Any]:
        if data is None or not path.startswith("$"):
            return False, None

        try:
            from jsonpath_ng.ext import parse
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "json_path_equals requires jsonpath-ng. Install dependencies with "
                "python3 -m pip install -r requirements.txt"
            ) from exc

        matches = [match.value for match in parse(path).find(data)]
        if not matches:
            return False, None

        actual = matches[0] if len(matches) == 1 else matches
        return True, actual

    def _values_equal(self, actual: Any, expected: Any) -> bool:
        if isinstance(actual, (dict, list)) or isinstance(expected, (dict, list)):
            return actual == expected
        return str(actual) == str(expected)


def run_all_checks(checks: List[SuccessCheck], context: Dict[str, Any] = None) -> List[VerificationResult]:
    runner = CheckRunner()
    if context:
        for k, v in context.items():
            runner.set_context(k, v)
    return [runner.run(c) for c in checks]
